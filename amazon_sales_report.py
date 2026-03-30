# ──────────────────────────────────────────────────────────────────────────────
# DEMO VERSION
# Sanitized for portfolio use. Hardcoded file paths, credentials, and email
# addresses have been replaced with environment variable references.
# Configure via a .env file or environment variables — see README.md for the
# full list of required keys. Requires Amazon Vendor Central access, a NetSuite
# account, and local login helper modules to run.
# ──────────────────────────────────────────────────────────────────────────────

import sys
import os
import time
import glob
import shutil
import pandas as pd
import traceback
import atexit
import csv
import io
import math
import re
import tempfile
import win32com.client as win32
import numpy as np
from scipy.ndimage import zoom
import matplotlib.image as mpimg
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from pathlib import Path
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import logging

# Add the Utilities folder to Python path
UTILITIES_PATH = os.environ.get("UTILITIES_PATH", "./Utilities")
sys.path.insert(0, UTILITIES_PATH)

# Import your login functions
from amazon_login import *
from netsuite_login import *

week_metrics = {}

DOWNLOAD_DIR = os.environ.get("DOWNLOAD_DIR", "./downloads")
LOG_DIR = os.environ.get("LOG_DIR", "./logs")
os.makedirs(LOG_DIR, exist_ok=True)

log_filename = datetime.now().strftime("%Y-%m-%d_%H-%M")
log_path = os.path.join(LOG_DIR, f"Amazon_Report_{log_filename}.txt")

log_file = open(log_path, "w", encoding="utf-8")

original_print = print

def print(*args, sep=' ', end='\n', file=None, flush=False):
    original_print(*args, sep=sep, end=end, file=file, flush=flush)
    if file is None:
        original_print(*args, sep=sep, end=end, file=log_file, flush=True)

def close_log():
    log_file.write(f"\n--- Script finished at {datetime.now():%Y-%m-%d %H:%M:%S} ---\n")
    log_file.close()

atexit.register(close_log)

def get_business_week_info(target_date=None):
    if target_date is None:
        target_date = datetime.now().date()
    elif isinstance(target_date, str):
        target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
    elif isinstance(target_date, datetime):
        target_date = target_date.date()

    days_to_sunday = (target_date.weekday() + 1) % 7
    week_sunday = target_date - timedelta(days=days_to_sunday)
    week_end = week_sunday + timedelta(days=6)

    year = week_end.year
    jan1 = datetime(year, 1, 1).date()
    days_back = (jan1.weekday() + 1) % 7
    first_sunday = jan1 - timedelta(days=days_back)

    days_since = (week_sunday - first_sunday).days
    week_number = (days_since // 7) + 1

    return {
        "week_number": week_number,
        "week_start": week_sunday,
        "week_end": week_end,
        "year": year,
        "label": f"Week {week_number:02d} ({week_sunday:%b %d} - {week_end:%b %d}, {year})"
    }

def get_current_business_week():
    return get_business_week_info()

def get_last_four_weeks():
    today = datetime.now().date()
    days_to_saturday = (today.weekday() - 5) % 7
    if days_to_saturday == 0:
        days_to_saturday = 7
    last_saturday = today - timedelta(days=days_to_saturday)

    weeks = []
    for i in range(4):
        saturday = last_saturday - timedelta(weeks=i)
        sunday = saturday - timedelta(days=6)
        weeks.append((
            sunday.strftime('%Y-%m-%d'),
            saturday.strftime('%Y-%m-%d'),
            f"{sunday.strftime('%Y-%m-%d')}to{saturday.strftime('%Y-%m-%d')}"
        ))
    return weeks[::-1]

def get_start_date(p):
    try:
        date_part = p.stem.split('_')[1]
        return datetime.strptime(date_part, "%Y-%m-%d")
    except:
        return datetime(1900, 1, 1)

def download_weekly_reports(driver, download_dir):
    print("\n" + "="*60)
    print("FINDING MOST RECENT COMPLETED WEEKS WITH AVAILABLE DATA")
    print("="*60)

    weeks = get_last_four_weeks()  # Get current 4-week window (most recent first in list? Wait, per original logic)
    successful_weeks = []
    csv_downloaded = False  # Flag to ensure only most recent week's CSV is downloaded

    # Try from most recent week backwards
    for sunday, saturday, week_range in weeks[::-1]:  # Reverse to start with most recent
        week_info = get_business_week_info(datetime.strptime(sunday, "%Y-%m-%d"))
        week_num = week_info["week_number"]
        print(f"\nTrying Week {week_num:02d} | {sunday} to {saturday}")

        url = f"https://vendorcentral.amazon.com/retail-analytics/dashboard/sales?compare-prior=true&compare-yoy=true&recurrence=year-to-date&submit=true&time-period=weekly&weekly-week={week_range}"
        driver.get(url)
        time.sleep(20)

        # Test if dashboard loaded successfully by trying to read metrics
        try:
            row_cells = WebDriverWait(driver, 15).until(
                EC.presence_of_all_elements_located((By.XPATH, "//table//tbody/tr[1]/td"))
            )
            revenue = row_cells[3].text.strip()
            prior_rev = row_cells[4].text.strip()
            yoy_rev = row_cells[5].text.strip()
            units = row_cells[6].text.strip()
            prior_units = row_cells[7].text.strip()
            yoy_units = row_cells[8].text.strip()

            # FIX: Check if revenue is a valid dollar amount before proceeding
            try:
                float(revenue.replace('$', '').replace(',', '').strip())
            except ValueError:
                print(f"   Invalid revenue format: '{revenue}' — skipping this week")
                continue

            print(f"   Dashboard loaded successfully: Revenue = {revenue}")

            # Capture metrics immediately (for all weeks)
            week_metrics[week_range] = {
                "Revenue": revenue,
                "Prior Period %": prior_rev,
                "YoY %": yoy_rev,
                "Last Week Units": units,
                "Prior Units %": prior_units,
                "YoY Units %": yoy_units,
                "Week Start": sunday,
                "Week End": saturday
            }
            successful_weeks.append((sunday, saturday, week_range))

        except Exception as e:
            print(f"   Dashboard metrics not available yet — skipping this week")
            continue

        # Proceed with CSV download ONLY for the most recent successful week
        if not csv_downloaded:
            try:
                csv_button = driver.execute_script("""
                    const btn = document.querySelector('kat-button#raw_csv_btn') 
                             || document.querySelector('kat-button[label="CSV"]');
                    return btn?.shadowRoot?.querySelector('button') || null;
                """)
                if not csv_button:
                    raise Exception("CSV button not found")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", csv_button)
                driver.execute_script("arguments[0].click();", csv_button)
                print("   Clicked CSV button")

                downloads_link = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[contains(., 'View and manage your downloads')]"))
                )
                driver.execute_script("arguments[0].click();", downloads_link)
                print("   Clicked 'View and manage your downloads'")

                time.sleep(5)

                start_dt = datetime.strptime(sunday, "%Y-%m-%d")
                end_dt = datetime.strptime(saturday, "%Y-%m-%d")
                start_str = f"{start_dt.month}-{start_dt.day}-{start_dt.year}"
                end_str = f"{end_dt.month}-{end_dt.day}-{end_dt.year}"
                filename_part = f"{start_str}_{end_str}"

                print(f"   Looking for filename containing: {filename_part}")

                # Wait for download link
                download_link = None
                for attempt in range(60):
                    try:
                        download_link = driver.find_element(
                            By.XPATH,
                            f"//kat-table-cell//a[contains(@href, 'Sales_ASIN_Manufacturing_Retail') and contains(@href, '{filename_part}') and normalize-space()='Download']"
                        )
                        if download_link.is_displayed():
                            print(f"   REPORT READY after {attempt * 4 + 4}s → clicking Download")
                            break
                    except:
                        pass
                    print(f"   Still processing... ({attempt * 4 + 4}s elapsed)")
                    time.sleep(4)
                else:
                    raise TimeoutError(f"Report never showed 'Download'")

                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", download_link)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", download_link)
                print("   File download started")

                # Wait for file
                downloaded_file = None
                start_time = time.time()
                while time.time() - start_time < 120:
                    for file_path in Path(download_dir).glob("*.csv"):
                        if filename_part in file_path.name and "crdownload" not in file_path.name:
                            downloaded_file = str(file_path)
                            break
                    if downloaded_file:
                        break
                    time.sleep(1)

                if not downloaded_file:
                    raise TimeoutError(f"File never appeared")

                # Rename file
                try:
                    date_part = downloaded_file.split("Weekly_")[1].split(".csv")[0]
                    start_str, end_str = date_part.split("_")
                    start_dt = datetime.strptime(start_str, "%m-%d-%Y").date()
                    end_dt = datetime.strptime(end_str, "%m-%d-%Y").date()
                    proper_start = start_dt.strftime("%Y-%m-%d")
                    proper_end = end_dt.strftime("%Y-%m-%d")

                    final_name = os.path.join(
                        download_dir,
                        f"W{week_num:02d}_{proper_start}_to_{proper_end}_Amazon_Sales.csv"
                    )
                except:
                    final_name = os.path.join(download_dir, f"Amazon_Sales_{week_range.replace('to', '_to_')}.csv")

                if downloaded_file != final_name:
                    os.replace(downloaded_file, final_name)

                print(f"   SUCCESS → {os.path.basename(final_name)}")
                csv_downloaded = True  # Mark as downloaded (only once)

            except Exception as e:
                print(f"   CSV download failed for most recent week: {e}")
                # Continue to capture metrics for older weeks anyway

        # No break here anymore — continue for all 4 weeks' metrics
        # (We can add a check if len(successful_weeks) == 4: break  but since only 4 weeks, loop ends naturally)

    print(f"\nSuccessfully captured metrics for {len(successful_weeks)} weeks (CSV only for most recent)")
    return successful_weeks


def download_ytd_report(driver, download_dir):
    ytd_url = "https://vendorcentral.amazon.com/retail-analytics/dashboard/sales?compare-prior=true&compare-yoy=true&recurrence=year-to-date&submit=true&time-period=recurring"


    print("\n" + "="*50)
    print("DOWNLOADING YEAR-TO-DATE REPORT")
    print("="*50)

    driver.get(ytd_url)
    time.sleep(10)

    try:
        row_cells = WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.XPATH, "//table//tbody/tr[1]/td"))
        )
        ytd_sales = row_cells[3].text.strip()
        ytd_yoy_sales = row_cells[4].text.strip()
        ytd_units = row_cells[5].text.strip()
        ytd_yoy_units = row_cells[6].text.strip()

        print(f"   YTD Ordered Sales    : {ytd_sales}")
        print(f"   YTD YoY Sales Var    : {ytd_yoy_sales}")
        print(f"   YTD Ordered Units    : {ytd_units}")
        print(f"   YTD YoY Units Var    : {ytd_yoy_units}")

        ytd_metrics = {
            "YTD Sales": ytd_sales,
            "YTD YoY Sales %": ytd_yoy_sales,
            "YTD Units": ytd_units,
            "YTD YoY Units %": ytd_yoy_units,
            "Period": "YTD"
        }
    except Exception as e:
        print(f"   Warning: Could not read YTD metrics: {e}")
        ytd_metrics = {"Period": "YTD"}

    try:
        csv_button = driver.execute_script("""
            const btn = document.querySelector('kat-button#raw_csv_btn') || document.querySelector('kat-button[label="CSV"]');
            return btn?.shadowRoot?.querySelector('button') || null;
        """)
        if not csv_button:
            raise Exception("CSV button not found")
        driver.execute_script("arguments[0].click();", csv_button)
        print("   Clicked YTD CSV button")

        downloads_link = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(., 'View and manage your downloads')]"))
        )
        driver.execute_script("arguments[0].click();", downloads_link)
        print("   Opened downloads panel")
        time.sleep(3)

        download_link = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, "//kat-table-cell//a[contains(@href, 'Todate') and text()='Download']"))
        )
        driver.execute_script("arguments[0].click();", download_link)
        print("   YTD download triggered")

        print("   Waiting for YTD file to appear...")
        downloaded_file = None
        final_name = None

        for _ in range(180):
            for file_path in Path(download_dir).glob("Sales_ASIN_Manufacturing_Retail_*Todate*.csv"):
                if "crdownload" in file_path.name:
                    continue
                downloaded_file = str(file_path)
                filename = file_path.name

                try:
                    date_part = filename.split("Todate_")[1].split(".csv")[0]
                    start_str, end_str = date_part.split("_")
                    start_dt = datetime.strptime(start_str, "%m-%d-%Y")
                    end_dt = datetime.strptime(end_str, "%m-%d-%Y")

                    proper_start = start_dt.strftime("%Y-%m-%d")
                    proper_end = end_dt.strftime("%Y-%m-%d")

                    final_name = os.path.join(
                        download_dir,
                        f"Amazon_Sales_YTD_{proper_start}_to_{proper_end}.csv"
                    )

                    ytd_metrics["Period"] = f"YTD {end_dt.year} (Jan 1 – {end_dt:%b %d})"
                    print(f"   YTD file found: {filename}")
                    print(f"   Renaming to: {os.path.basename(final_name)}")
                    break
                except Exception as e:
                    print(f"   Warning: Could not parse YTD filename: {e}")
                    continue

            if final_name:
                break
            time.sleep(2)

        if downloaded_file and final_name and downloaded_file != final_name:
            try:
                os.replace(downloaded_file, final_name)
                print(f"   YTD SUCCESS → {os.path.basename(final_name)}")
            except Exception as e:
                print(f"   Rename failed: {e}")
                print(f"   Keeping → {os.path.basename(downloaded_file)}")
        elif downloaded_file:
            print(f"   YTD SUCCESS → {os.path.basename(downloaded_file)}")
        else:
            print("   YTD file not detected — continuing with metrics only")

        return ytd_metrics

    except Exception as e:
        print(f"   YTD download failed: {e}")
        driver.save_screenshot("ERROR_YTD.png")
        return ytd_metrics


def download_traffic_report(driver, download_dir, week_range, week_num):
    traffic_url = f"https://vendorcentral.amazon.com/retail-analytics/dashboard/traffic?compare-prior=true&compare-yoy=true&submit=true&time-period=weekly&weekly-week={week_range}"

    print("\n" + "="*60)
    print(f"DOWNLOADING TRAFFIC REPORT - Week {week_num}")
    print("="*60)

    driver.get(traffic_url)
    time.sleep(15)

    # ====================== NEW: EXTRACT TRAFFIC TOTALS ======================
    traffic_metrics = {
        "Glance Views": None,
        "TY 2wks Ago Views": None,
        "YoY VAR Views": None
    }

    try:
        row_cells = WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.XPATH, "//table//tbody/tr[1]/td"))
        )

        # Your exact columns (0-based index)
        traffic_metrics["Glance Views"]     = row_cells[3].text.strip()   # td[4]
        traffic_metrics["TY 2wks Ago Views"] = row_cells[4].text.strip()  # td[5]
        traffic_metrics["YoY VAR Views"]    = row_cells[5].text.strip()   # td[6]

        print(f"   Glance Views (Total)     : {traffic_metrics['Glance Views']}")
        print(f"   TY 2wks Ago Views        : {traffic_metrics['TY 2wks Ago Views']}")
        print(f"   YoY VAR Views            : {traffic_metrics['YoY VAR Views']}")

    except Exception as e:
        print(f"   Warning: Could not read Traffic totals: {e}")
        driver.save_screenshot(f"ERROR_Traffic_Totals_W{week_num}.png")
    # =========================================================================

    # ==================== KEEP ALL YOUR ORIGINAL CSV DOWNLOAD CODE ====================
    try:
        csv_button = driver.execute_script("""
            const btn = document.querySelector('kat-button#raw_csv_btn') 
                     || document.querySelector('kat-button[label="CSV"]');
            return btn?.shadowRoot?.querySelector('button') || null;
        """)
        if csv_button:
            driver.execute_script("arguments[0].click();", csv_button)
            print("   Clicked Traffic CSV button")
        else:
            raise Exception("CSV button not found for Traffic")

        downloads_link = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(., 'View and manage your downloads')]"))
        )
        driver.execute_script("arguments[0].click();", downloads_link)

        start_dt = datetime.strptime(week_range.split("to")[0], "%Y-%m-%d")
        end_dt = datetime.strptime(week_range.split("to")[1], "%Y-%m-%d")
        start_str = f"{start_dt.month}-{start_dt.day}-{start_dt.year}"
        end_str = f"{end_dt.month}-{end_dt.day}-{end_dt.year}"
        filename_part = f"{start_str}_{end_str}"

        print(f"   Looking for filename containing: {filename_part}")

        download_link = None
        for attempt in range(60):
            try:
                download_link = driver.find_element(
                    By.XPATH,
                    f"//kat-table-cell//a[contains(@href, 'Traffic_ASIN') and contains(@href, '{filename_part}') and normalize-space()='Download']"
                )
                if download_link.is_displayed():
                    print(f"   REPORT READY after {attempt * 4 + 4}s → clicking Download")
                    break
            except:
                pass
            print(f"   Still processing... ({attempt * 4 + 4}s elapsed)")
            time.sleep(4)
        else:
            raise TimeoutError(f"Traffic report never showed 'Download'")

        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", download_link)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", download_link)
        print("   Traffic download started")

        downloaded_file = None
        start_time = time.time()
        while time.time() - start_time < 120:
            for file_path in Path(download_dir).glob("*.csv"):
                if "Traffic_ASIN" in file_path.name and filename_part in file_path.name and "crdownload" not in file_path.name:
                    downloaded_file = str(file_path)
                    break
            if downloaded_file:
                break
            time.sleep(1)

        if downloaded_file:
            proper_start = start_dt.strftime("%Y-%m-%d")
            proper_end = end_dt.strftime("%Y-%m-%d")
            final_name = os.path.join(download_dir, f"W{week_num:02d}_{proper_start}_to_{proper_end}_Amazon_Traffic.csv")
            os.replace(downloaded_file, final_name)
            print(f"   SUCCESS → {os.path.basename(final_name)}")
        else:
            print("   Traffic file not detected")

    except Exception as e:
        print(f"   Traffic report failed: {e}")
        driver.save_screenshot(f"ERROR_TRAFFIC_W{week_num}.png")

    return traffic_metrics   # ← Return the three values


def download_inventory_report(driver, download_dir, week_range, week_num):
    inventory_url = f"https://vendorcentral.amazon.com/retail-analytics/dashboard/inventory?submit=true&time-period=weekly&weekly-week={week_range}"

    print("\n" + "="*60)
    print(f"DOWNLOADING INVENTORY REPORT - Week {week_num}")
    print("="*60)

    driver.get(inventory_url)
    time.sleep(15)

    # ====================== EXTRACT INVENTORY TOTALS (Dashboard Grand Totals) ======================
    inventory_metrics = {
        "Sell-Thru %": None,
        "Open PO QTY": None,
        "On Hand INV $": None,
        "On Hand Units": None
    }

    try:
        row_cells = WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.XPATH, "//table//tbody/tr[1]/td"))
        )

        # Your exact columns (0-based index)
        inventory_metrics["Open PO QTY"]   = row_cells[7].text.strip()    # td[8]
        inventory_metrics["On Hand INV $"] = row_cells[13].text.strip()   # td[14]
        inventory_metrics["On Hand Units"] = row_cells[14].text.strip()   # td[15]
        inventory_metrics["Sell-Thru %"]   = row_cells[17].text.strip()   # td[18]

        print(f"   Sell-Thru % (Total)      : {inventory_metrics['Sell-Thru %']}")
        print(f"   Open PO QTY (Total)      : {inventory_metrics['Open PO QTY']}")
        print(f"   On Hand INV $ (Total)    : {inventory_metrics['On Hand INV $']}")
        print(f"   On Hand Units (Total)    : {inventory_metrics['On Hand Units']}")

    except Exception as e:
        print(f"   Warning: Could not read Inventory totals: {e}")
        driver.save_screenshot(f"ERROR_Inventory_Totals_W{week_num}.png")
    # ============================================================================================

    # ==================== KEEP ALL YOUR ORIGINAL CSV DOWNLOAD CODE ====================
    try:
        csv_button = driver.execute_script("""
            const btn = document.querySelector('kat-button#raw_csv_btn') 
                     || document.querySelector('kat-button[label="CSV"]');
            return btn?.shadowRoot?.querySelector('button') || null;
        """)
        if csv_button:
            driver.execute_script("arguments[0].click();", csv_button)
            print("   Clicked Inventory CSV button")
        else:
            raise Exception("CSV button not found for Inventory")

        downloads_link = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, "//a[contains(., 'View and manage your downloads')]"))
        )
        driver.execute_script("arguments[0].click();", downloads_link)

        start_dt = datetime.strptime(week_range.split("to")[0], "%Y-%m-%d")
        end_dt = datetime.strptime(week_range.split("to")[1], "%Y-%m-%d")
        start_str = f"{start_dt.month}-{start_dt.day}-{start_dt.year}"
        end_str = f"{end_dt.month}-{end_dt.day}-{end_dt.year}"
        filename_part = f"{start_str}_{end_str}"

        print(f"   Looking for filename containing: {filename_part}")

        download_link = None
        for attempt in range(60):
            try:
                download_link = driver.find_element(
                    By.XPATH,
                    f"//kat-table-cell//a[contains(@href, 'Inventory_ASIN_Manufacturing_Retail') and contains(@href, '{filename_part}') and normalize-space()='Download']"
                )
                if download_link.is_displayed():
                    print(f"   REPORT READY after {attempt * 4 + 4}s → clicking Download")
                    break
            except:
                pass
            print(f"   Still processing... ({attempt * 4 + 4}s elapsed)")
            time.sleep(4)
        else:
            raise TimeoutError(f"Inventory report never showed 'Download'")

        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", download_link)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", download_link)
        print("   Inventory download started")

        downloaded_file = None
        start_time = time.time()
        while time.time() - start_time < 120:
            for file_path in Path(download_dir).glob("*.csv"):
                if "Inventory_ASIN_Manufacturing_Retail" in file_path.name and filename_part in file_path.name and "crdownload" not in file_path.name:
                    downloaded_file = str(file_path)
                    break
            if downloaded_file:
                break
            time.sleep(1)

        if downloaded_file:
            proper_start = start_dt.strftime("%Y-%m-%d")
            proper_end = end_dt.strftime("%Y-%m-%d")
            final_name = os.path.join(download_dir, f"W{week_num:02d}_{proper_start}_to_{proper_end}_Amazon_Inventory.csv")
            os.replace(downloaded_file, final_name)
            print(f"   SUCCESS → {os.path.basename(final_name)}")
        else:
            print("   Inventory file not detected")

    except Exception as e:
        print(f"   Inventory report failed: {e}")
        driver.save_screenshot(f"ERROR_INVENTORY_W{week_num}.png")

    return inventory_metrics   # ← Now returns 4 totals


def download_business_metrics(driver, week_range):
    print("\n" + "=" * 60)
    print("FINDING MOST RECENT WEEK BUSINESS DATA")
    print("=" * 60)


    url = f"https://vendorcentral.amazon.com/retail-analytics/dashboard/sales?compare-yoy=true&programView=business&submit=true&time-period=weekly&weekly-week={week_range}"
    driver.get(url)
    time.sleep(10)
    try:
        rev_elem = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div/div/div/div[2]/div[3]/div[1]/div/table/tbody/tr[1]/td[4]"))
        )
        yoy_elem = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div/div/div[2]/div[3]/div[1]/div/table/tbody/tr[1]/td[5]")
        rev = rev_elem.text.strip()
        yoy = yoy_elem.text.strip()
        print(f'Revenue: {rev}, YoY: {yoy}')
        return {"Revenue": rev, "YoY": yoy}
    except Exception as e:
        print(f"   Failed to fetch business metrics: {e}")
        return None


def close_excel_workbook(file_path):
    try:
        excel = win32.GetActiveObject('Excel.Application')
    except:
        excel = win32.Dispatch('Excel.Application')
    closed = False
    for wb in excel.Workbooks:
        if os.path.samefile(os.path.abspath(wb.FullName), os.path.abspath(file_path)):
            wb.Close(SaveChanges=False)
            closed = True
            break
    if closed:
        print(f"Closed open instance of {os.path.basename(file_path)} in Excel")
    excel.Quit()
    del excel
    time.sleep(1)  # Allow process to exit

def clean_numeric_keep_na(series):
    """
    Convert column to numeric while preserving original NaN/missing values.
    This replaces the old clean_numeric_keep_na() that had .fillna(0).
    """
    if series is None or len(series) == 0:
        return pd.Series([np.nan] * len(series))
    s = series.astype(str).str.replace(r'[\$,%]', '', regex=True).str.strip()
    return pd.to_numeric(s, errors='coerce')  # ← NO .fillna(0) here


def format_for_display(value, col_name=""):
    """
    Final display formatting — ONLY used when building HTML tables for email.
    This is where NaN finally becomes '' (empty string).
    """
    if pd.isna(value):
        return ""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return str(value)

    # Explicit column to format mapping
    if col_name in ['WTD Ordered SLS $', 'YTD Ordered SLS $', 'On Hand INV $']:
        return f"${v:,.0f}"
    elif col_name in ['ASP $', 'COGS', 'Margin ($)']:
        return f"${v:,.2f}"
    elif col_name in ['TY 2wks Ago SLS %', 'YoY VAR SLS', 'TY 2wks Ago UNITS', 'YoY VAR UNITS',
                      'YoY VAR YTD SLS', 'YoY VAR YTD UNITS', 'Sell-Thru %', 'TY 2wks Ago Views',
                      'YoY VAR Views', 'Conversion Rate', 'Margin (%)']:
        return f"{v * 100 :.1f}%"
    elif col_name in ['WTD Ordered UNITS', 'YTD Ordered UNITS', 'Open PO QTY', 'On Hand Units',
                      'Glance Views', 'Rank', 'Wks of INV']:
        return f"{int(v):,}" if v == int(v) else f"{v:,.0f}"
    else:
        # Fallback for strings or other
        return str(value)

def main():
    start_time = time.time()  # Start timer

    # Clear old Amazon files (keep NetSuite-related files)
    for pattern in ["Amazon_Sales*.csv", "Sales_ASIN_Manufacturing_Retail_*.csv"]:
        for f in Path(DOWNLOAD_DIR).glob(pattern):
            try:
                f.unlink()
            except:
                pass

    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service as ChromeService
    from webdriver_manager.chrome import ChromeDriverManager

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--start-maximized")
    prefs = {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    chrome_options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)

    def log(msg):
        print(f"[{time.strftime('%H:%M:%S')}] {msg}")

    try:
        driver.get("https://vendorcentral.amazon.com")
        log("Starting Amazon login...")
        amazon_login(driver, log)
        log("Amazon login successful! Starting report downloads...")

        current_week = get_current_business_week()
        print("\n" + "=" * 60)
        print("          AMAZON BUSINESS WEEK DETECTION")
        print("=" * 60)
        print(f"Today is: {datetime.now().strftime('%A, %B %d, %Y')}")
        print(f"Current Business Week: {current_week['label']}")
        print(f"Downloading last 4 completed weeks...")
        print("=" * 60 + "\n")

        successful_weeks = download_weekly_reports(driver, DOWNLOAD_DIR)
        log("All 4 weekly reports capture")

        # === DETERMINE MOST RECENT COMPLETED WEEK (for subject line and any titles) ===
        # successful_weeks[0] is the most recent completed week
        most_recent_sunday_str = successful_weeks[0][0]  # e.g. '2026-02-01'
        most_recent_week_info = get_business_week_info(datetime.strptime(most_recent_sunday_str, "%Y-%m-%d"))
        week_num_for_email = most_recent_week_info["week_number"]
        year_for_email = most_recent_week_info["year"]

        print(f"Email will use: Year {year_for_email} - Week {week_num_for_email:02d}")

        # === DOWNLOAD TRAFFIC & INVENTORY FOR MOST RECENT WEEK ONLY ===
        weeks = get_last_four_weeks()
        most_recent_week_range = weeks[-1][2]  # e.g. "2026-01-11to2026-01-17"
        week_info = get_business_week_info(datetime.strptime(weeks[-1][0], "%Y-%m-%d"))
        week_num = week_info["week_number"]

        print(f"\nDownloading Traffic and Inventory reports for most recent week: Week {week_num}")

        traffic_metrics = download_traffic_report(driver, DOWNLOAD_DIR, most_recent_week_range, week_num)
        inventory_metrics = download_inventory_report(driver, DOWNLOAD_DIR, most_recent_week_range, week_num)

        ytd_metrics = download_ytd_report(driver, DOWNLOAD_DIR)
        log("YTD report processing complete.")

        business_metrics = download_business_metrics(driver, most_recent_week_range)

        # Amazon Executive Summary (unchanged)
        print("\n" + "=" * 90)
        print("                    AMAZON SALES EXECUTIVE SUMMARY")
        print("=" * 90)

        sorted_weeks = sorted(week_metrics.items(), key=lambda x: x[1]["Week Start"])

        revenues = []
        print(f"{'Week':<30} {'Revenue':>14}  {'Prior Period':>12}  {'YoY':>10}  {'Units':>14}")
        print("-" * 90)

        for week_range, data in sorted_weeks:
            try:
                rev_clean = data["Revenue"].replace("$", "").replace(",", "").strip()
                rev_float = float(rev_clean)
                rev_rounded = round(rev_float)
                revenues.append(rev_float)
            except:
                rev_rounded = "ERROR"
                revenues.append(0)

            prior_str = data["Prior Period %"].rstrip("%").strip()
            prior_str = f"{float(prior_str):+.1f}%" if prior_str.replace('-','').replace('.','').replace('+','').isdigit() else prior_str

            yoy_str = data["YoY %"].rstrip("%").strip()
            yoy_str = f"{float(yoy_str):+.1f}%" if yoy_str.replace('-','').replace('.','').replace('+','').isdigit() else yoy_str

            units_raw = data["Last Week Units"].replace(",", "").strip()
            units = f"{int(units_raw):,}" if units_raw.isdigit() else "N/A"

            start_dt = datetime.strptime(data["Week Start"], "%Y-%m-%d").date()
            end_dt = datetime.strptime(data["Week End"], "%Y-%m-%d").date()

            week_info = get_business_week_info(start_dt)
            week_num = week_info["week_number"]
            week_label = f"Week {week_num:02d} | {start_dt:%Y-%m-%d} to {end_dt:%Y-%m-%d}"

            print(f"{week_label:<30} ${rev_rounded:>12,}  {prior_str:>12}  {yoy_str:>10}  {units:>10}")

        print("\n" + "-" * 90)
        print(f"{ytd_metrics.get('Period', 'YEAR-TO-DATE')}")
        print("-" * 90)

        if "YTD Sales" in ytd_metrics:
            try:
                ytd_sales_clean = ytd_metrics["YTD Sales"].replace("$", "").replace(",", "").strip()
                ytd_sales_float = float(ytd_sales_clean)
                ytd_sales_rounded = round(ytd_sales_float)

                ytd_yoy_sales = ytd_metrics["YTD YoY Sales %"].rstrip("%").strip()
                yoy_sales_str = f"{float(ytd_yoy_sales):+.1f}%"

                ytd_units_clean = ytd_metrics["YTD Units"].replace(",", "").strip()
                ytd_units_int = int(float(ytd_units_clean))

                ytd_yoy_units = ytd_metrics["YTD YoY Units %"].rstrip("%").strip()
                yoy_units_str = f"{float(ytd_yoy_units):+.1f}%"

                print(f"{'YTD Ordered Sales':<30} ${ytd_sales_rounded:>12,}")
                print(f"{'YTD YoY Sales Growth':<30} {yoy_sales_str:>12}")
                print(f"{'YTD Ordered Units':<30} {ytd_units_int:>12,}")
                print(f"{'YTD YoY Units Growth':<30} {yoy_units_str:>12}")
            except:
                print("  Could not parse YTD numbers")

        if len(revenues) == 4 and sum(revenues) > 0:
            four_week_avg = sum(revenues) / 4
            most_recent = revenues[-1]
            growth_vs_avg = (most_recent / four_week_avg - 1) * 100

            print("-" * 90)
            print(f"{'4-Week Average Sales':<30} ${round(four_week_avg):>12,}")
            print(f"{'Most Recent Week Sales':<30} ${round(most_recent):>12,}")
            print(f"{'Growth vs 4-Week Average':<30} {growth_vs_avg:+.1f}%")
            print("=" * 90)
        else:
            print("Could not calculate summary — missing data")
            print("=" * 90)

        # Update YOY Weekly Revenue Data.xlsx with most recent week's revenue
        if sorted_weeks:
            week_range, data = sorted_weeks[-1]
            revenue = data["Revenue"]
            revenue_clean = revenue.replace("$", "").replace(",", "").strip()
            revenue_float = float(revenue_clean)
            week_start = data["Week Start"]
            week_info = get_business_week_info(week_start)
            week_num = week_info["week_number"]
            year = week_info["year"]

            yoy_path = os.path.join(os.environ.get("SALES_DATA_DIR", "./data"), "YOY Weekly Revenue Data.xlsx"
            from openpyxl import load_workbook
            from openpyxl.chart import LineChart, Reference
            from openpyxl.chart.axis import ChartLines
            from openpyxl.styles import Font
            close_excel_workbook(yoy_path)
            wb = load_workbook(yoy_path)
            ws = wb.active

            accounting_fmt = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

            # Set font for all cells
            font = Font(name='Calibri', size=11)
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font

            # Find or add year column
            year_header = f"{year} SALES $"
            year_col = None
            for col in range(1, ws.max_column + 1):
                header_val = ws.cell(1, col).value
                if header_val and str(header_val).strip() == year_header:
                    year_col = col
                    break
            if year_col is None:
                year_col = ws.max_column + 1
                ws.cell(1, year_col).value = year_header

            # Find or add week row
            week_header = str(week_num)
            week_row = None
            for row in range(2, ws.max_row + 2):
                val = ws.cell(row, 1).value
                if val is not None and str(val).strip() == week_header:
                    week_row = row
                    break
            if week_row is None:
                week_row = ws.max_row + 1
                ws.cell(week_row, 1).value = int(week_num)

            # Set value and format
            cell = ws.cell(week_row, year_col)
            cell.value = revenue_float
            cell.number_format = accounting_fmt

            # Add or update line chart
            ws._charts = []  # Remove existing charts
            chart = LineChart()
            chart.title = None
            chart.x_axis.title = None
            chart.y_axis.title = None

            # Force Y-axis to start at exactly $150,000
            chart.y_axis.scaling.min = 125000
            chart.y_axis.scaling.max = None  # let Excel auto-scale the top
            chart.y_axis.scaling.orientation = "minMax"

            # Make sure both primary axes are visible
            chart.x_axis.delete = False
            chart.y_axis.delete = False

            # Major + Minor gridlines on both axes
            chart.y_axis.majorGridlines = ChartLines()
            chart.y_axis.minorGridlines = ChartLines()
            chart.x_axis.majorGridlines = ChartLines()
            chart.x_axis.minorGridlines = ChartLines()

            # Legend at the top, without overlapping the chart
            chart.legend.position = 't'  # top
            chart.legend.overlay = False

            # Size (large enough so legend doesn't overlap lines)
            chart.height = 15
            chart.width = 24

            # Data range = last 4 full years + current year up to the latest week that has data
            min_col = max(2, year_col - 3)
            max_row = 1
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, year_col).value is not None:
                    max_row = r

            data = Reference(ws, min_col=min_col, min_row=1, max_col=year_col, max_row=max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws.add_chart(chart, "H2")

            wb.save(yoy_path)
            close_excel_workbook(yoy_path)
            print(f"Updated YOY file with Week {week_num} {year} revenue: ${revenue_float:,.0f}")

            # ===============================================================
            # Generate YoY line graph in Python (after Excel is closed)
            # ===============================================================
            print("Generating YoY line graph with Matplotlib...")

            df_yoy = pd.read_excel(yoy_path)

            # Column 0 = Week numbers (1-52)
            weeks = df_yoy.iloc[:, 0].dropna().astype(int).values

            # Find current year column
            year_header = f"{year} SALES $"
            if year_header not in df_yoy.columns:
                raise ValueError(f"Current year column '{year_header}' not found")

            current_col_idx = df_yoy.columns.get_loc(year_header)

            # Get up to 3 previous + current (4 total; handle if fewer)
            min_col_idx = max(1, current_col_idx - 3)
            year_cols = df_yoy.columns[min_col_idx: current_col_idx + 1]

            years_data = df_yoy[year_cols]

            fig, ax = plt.subplots(figsize=(12, 7.5))
            fig.patch.set_edgecolor('black')
            fig.patch.set_linewidth(2)

            colors = plt.cm.tab10(np.linspace(0, 1, len(year_cols)))

            last_points = []

            for i, col in enumerate(year_cols):
                data = years_data[col].values[:week_num]  # Limit all years to current week
                valid_mask = ~pd.isna(data)
                valid_weeks = weeks[:len(data)][valid_mask]  # Slice weeks to match data length
                valid_data = data[valid_mask]

                color = colors[i]

                # Depth effect: multiple offset shadow lines
                for offset in [3, 2, 1]:  # Multi-layer shadows
                    ax.plot(valid_weeks, valid_data - offset * 1000, color='black', alpha=0.1 / offset,
                            linewidth=4 + offset, zorder=1)

                # Main line (thicker, on top)
                marker = 'o' if col == year_header else None
                ax.plot(valid_weeks, valid_data, label=col, color=color, marker=marker,
                        linewidth=3.5, zorder=2, alpha=0.95)

                # Collect last point for labeling later
                if len(valid_data) > 0:
                    last_week = valid_weeks[-1]
                    last_value = valid_data[-1]
                    last_points.append({'col': col, 'week': last_week, 'value': last_value, 'color': color})

            # Sort last_points by value descending
            last_points.sort(key=lambda x: x['value'], reverse=True)

            # Determine threshold for closeness (2% of max value)
            if last_points:
                max_value = max(p['value'] for p in last_points)
                threshold = 0.02 * max_value
            else:
                threshold = 10000  # Fallback

            # Place labels with staggered x_offsets if close
            previous_y = None
            x_offset = 0.2
            step = 0.8  # Additional offset if close

            for point in last_points:
                y = point['value']
                current_x_offset = x_offset
                if previous_y is not None and abs(previous_y - y) < threshold:
                    current_x_offset += step
                ax.text(point['week'] + current_x_offset, y, f'${int(y):,}',
                        color='black', fontsize=10, ha='left', va='center', fontweight='bold')
                previous_y = y

            ax.set_title('YoY Weekly Revenue', fontsize=14, fontweight='bold')
            ax.set_xlabel('Week Number')
            ax.set_ylabel('Sales $')
            ax.set_ylim(bottom=125000)
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'${int(x):,}' if not math.isnan(x) else ''))
            ax.grid(True, linestyle='--', alpha=0.7)
            ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.12), ncol=5, frameon=True)
            ax.set_xticks(range(1, week_num + 1, 2))
            ax.set_xlim(1, week_num + 4)  # Extended to make space for staggered labels

            chart_img_path = os.path.join(os.environ.get('SALES_DATA_DIR', './data'), 'PNG Files\YoY Line Graph.png'
            plt.savefig(chart_img_path, format='png', bbox_inches='tight')
            plt.close(fig)

            print(f"   YoY line graph generated → {chart_img_path}")

            # Resize to ~half size
            img = mpimg.imread(chart_img_path)
            resized_img = zoom(img, (0.90, 0.90, 1) if len(img.shape) == 3 else (0.90, 0.90))
            resized_img = np.clip(resized_img, 0.0, 1.0)

            resized_img_path = os.path.join(os.environ.get('SALES_DATA_DIR', './data'), 'PNG Files\resized_yoy_graph.png'
            mpimg.imsave(resized_img_path, resized_img)


        # ────────────────────────────────────────────────
        #          MOVE FILES TO FINAL FOLDER
        # ────────────────────────────────────────────────
        FINAL_DATA_DIR = os.environ.get("SALES_DATA_DIR", "./data")
        os.makedirs(FINAL_DATA_DIR, exist_ok=True)

        print("\n" + "-" * 60)
        print("CLEARING OLD FILES FROM FINAL FOLDER")
        print("-" * 60)
        cleared = 0
        skip_file = "HardistyAmazonRecentItems_Processed.csv"
        for f in Path(FINAL_DATA_DIR).glob("*.csv"):
            if f.name == skip_file:
                print(f"   Skipping → {f.name}")
                continue
            try:
                f.unlink()
                print(f"   Deleted old → {f.name}")
                cleared += 1
            except:
                pass
        print(f"   Cleared {cleared} old files" if cleared else "   No old files found")

        print("\n" + "-" * 60)
        print("MOVING NEW REPORTS TO FINAL FOLDER")
        print("-" * 60)
        moved = 0

        for src in Path(DOWNLOAD_DIR).glob("W*_Amazon_Sales.csv"):
            dest = Path(FINAL_DATA_DIR) / src.name
            try:
                os.replace(src, dest)
                print(f"   Moved → {src.name}")
                moved += 1
            except Exception as e:
                print(f"   Failed to move {src.name}: {e}")

        for src in Path(DOWNLOAD_DIR).glob("Amazon_Sales_YTD_*.csv"):
            dest = Path(FINAL_DATA_DIR) / src.name
            try:
                os.replace(src, dest)
                print(f"   Moved → {src.name}")
                moved += 1
            except Exception as e:
                print(f"   Failed to move {src.name}: {e}")

        for src in Path(DOWNLOAD_DIR).glob("W*_Amazon_Traffic.csv"):
            dest = Path(FINAL_DATA_DIR) / src.name
            try:
                os.replace(src, dest)
                print(f"   Moved → {src.name}")
                moved += 1
            except Exception as e:
                print(f"   Failed to move {src.name}: {e}")

        for src in Path(DOWNLOAD_DIR).glob("W*_Amazon_Inventory.csv"):
            dest = Path(FINAL_DATA_DIR) / src.name
            try:
                os.replace(src, dest)
                print(f"   Moved → {src.name}")
                moved += 1
            except Exception as e:
                print(f"   Failed to move {src.name}: {e}")

        for src in Path(DOWNLOAD_DIR).glob("HardistyAmazonRecentItems_Processed.csv"):
            dest = Path(FINAL_DATA_DIR) / src.name
            try:
                os.replace(src, dest)
                print(f"   Moved → {src.name}")
                moved += 1
            except Exception as e:
                print(f"   Failed to move {src.name}: {e}")

        print(f"\nAll {moved} new reports moved to:\n   {FINAL_DATA_DIR}")
        print("=" * 90)

        # ────────────────────────────────────────────────
        #          POST-PROCESS LATEST WEEKLY + YTD + TRAFFIC + INVENTORY
        # ────────────────────────────────────────────────
        print("\n" + "=" * 90)
        print("          POST-PROCESSING WEEKLY + YTD + TRAFFIC + INVENTORY")
        print("=" * 90)

        processed_netsuite = os.path.join(FINAL_DATA_DIR, "HardistyAmazonRecentItems_Processed.csv")

        if not os.path.exists(processed_netsuite):
            print("   NetSuite processed file not found — skipping post-processing")
        else:
            df_lookup = pd.read_csv(processed_netsuite, dtype=str)
            print(f"   Loaded NetSuite lookup: {len(df_lookup)} rows")

            lookup_dict = {}
            for _, row in df_lookup.iterrows():
                cpn = str(row.get('CPN', '')).strip()
                if cpn:
                    lookup_dict[cpn] = (
                    row.get('Model', ''), row.get('SKUs', ''), row.get('Description', ''), row.get('Category', ''))

            # ── Define shared column lists here ──
            cols_to_clean = [
                'WTD Ordered SLS $', 'TY 2wks Ago SLS %', 'YoY VAR SLS',
                'WTD Ordered UNITS', 'TY 2wks Ago UNITS', 'YoY VAR UNITS',
                'YTD Ordered SLS $', 'YoY VAR YTD SLS', 'YTD Ordered UNITS',
                'YoY VAR YTD UNITS', 'Open PO QTY', 'On Hand INV $',
                'On Hand Units', 'Glance Views', 'TY 2wks Ago Views',
                'YoY VAR Views', 'Conversion Rate', 'ASP $', 'COGS',
                'Margin ($)', 'Margin (%)'
            ]

            def process_amazon_file(file_path, copy_name):
                copy_path = os.path.join(FINAL_DATA_DIR, copy_name)
                shutil.copy2(file_path, copy_path)
                print(f"   Copied to: {copy_name}")

                try:
                    with open(copy_path, 'r', encoding='utf-8-sig') as f:
                        lines = f.readlines()

                    header_idx = None
                    for idx, line in enumerate(lines):
                        stripped = line.strip()
                        if not stripped:
                            continue
                        # Strict check: first field after stripping quotes/spaces is exactly "ASIN"
                        first_field = stripped.split(',')[0].strip().strip('"\'')
                        if first_field == 'ASIN':
                            header_idx = idx
                            print(f"   Found strict ASIN header at line {idx + 1}: {stripped[:120]}...")
                            break

                    if header_idx is None:
                        print("   No strict ASIN header found — falling back to heuristic")
                        # Fallback: look for line with many columns containing "ASIN"
                        for idx, line in enumerate(lines):
                            stripped = line.strip()
                            if len(stripped.split(',')) > 8 and 'ASIN' in stripped[:100]:
                                header_idx = idx
                                print(f"   Found heuristic header at line {idx + 1}")
                                break

                    if header_idx is None:
                        print("   Could not find valid header — skipping")
                        return None

                    clean_csv = ''.join(lines[header_idx:])
                    df = pd.read_csv(io.StringIO(clean_csv), quoting=csv.QUOTE_ALL, dtype=str, on_bad_lines='warn')

                    print(f"   Loaded {len(df)} rows")
                    print(f"   First 6 columns: {list(df.columns[:6])}")
                    return df

                except Exception as e:
                    print(f"   Failed to process {copy_name}: {e}")
                    traceback.print_exc()
                    return None

            # Weekly (already processed)
            weekly_files = list(Path(FINAL_DATA_DIR).glob("W*_Amazon_Sales.csv"))
            df_weekly = None
            weekly_copy_path = None
            if weekly_files:
                latest_weekly = max(weekly_files, key=get_start_date)
                df_weekly = process_amazon_file(latest_weekly, "Latest_Weekly_Processed.csv")
                weekly_copy_path = os.path.join(FINAL_DATA_DIR, "Latest_Weekly_Processed.csv")

            # YTD (already processed)
            ytd_files = list(Path(FINAL_DATA_DIR).glob("Amazon_Sales_YTD_*.csv"))
            df_ytd = None
            if ytd_files:
                latest_ytd = max(ytd_files, key=lambda p: os.path.getmtime(p))
                df_ytd = process_amazon_file(latest_ytd, "Latest_YTD_Processed.csv")

            # Traffic
            traffic_files = list(Path(FINAL_DATA_DIR).glob("W*_Amazon_Traffic.csv"))
            df_traffic = None
            if traffic_files:
                latest_traffic = max(traffic_files, key=get_start_date)
                df_traffic = process_amazon_file(latest_traffic, "Latest_Traffic_Processed.csv")

            # Inventory
            inventory_files = list(Path(FINAL_DATA_DIR).glob("W*_Amazon_Inventory.csv"))
            df_inventory = None
            if inventory_files:
                latest_inventory = max(inventory_files, key=get_start_date)
                df_inventory = process_amazon_file(latest_inventory, "Latest_Inventory_Processed.csv")

            # === COLLECT UNIQUE ASINS FROM ALL REPORTS ===
            all_asins = set()

            def extract_asins(df):
                if df is not None:
                    asin_col = next((c for c in df.columns if 'asin' in str(c).lower()), None)
                    if asin_col:
                        return set(df[asin_col].astype(str).str.strip().str.upper().dropna())
                return set()

            all_asins.update(extract_asins(df_weekly))
            all_asins.update(extract_asins(df_ytd))
            all_asins.update(extract_asins(df_traffic))
            all_asins.update(extract_asins(df_inventory))

            if not all_asins:
                print("   No ASINs found in any report — skipping final report generation")
            else:
                print(f"   Found {len(all_asins)} unique ASINs across all reports")

                # Create master DF with unique ASINs
                df_master = pd.DataFrame({'ASIN': list(all_asins)})

                # Add Model/SKUs/Description/Category using lookup
                df_master['Model'] = ''
                df_master['SKUs'] = ''
                df_master['Description'] = ''
                df_master['Category'] = ''

                for idx, row in df_master.iterrows():
                    asin = row['ASIN']
                    if asin in lookup_dict:
                        model, skus, desc, cat = lookup_dict[asin]
                        df_master.at[idx, 'Model'] = model
                        df_master.at[idx, 'SKUs'] = skus
                        df_master.at[idx, 'Description'] = desc
                        cat_stripped = re.sub(r'^\([A-Za-z]+\)\s*', '', str(cat)).strip()
                        df_master.at[idx, 'Category'] = cat_stripped

                print("   Added lookup info to master DF")

                # Now merge data from each report
                def merge_report(master_df, report_df, suffixes, asin_col_name):
                    if report_df is not None:
                        asin_col = next((c for c in report_df.columns if 'asin' in str(c).lower()), None)
                        if asin_col:
                            report_df[asin_col] = report_df[asin_col].astype(str).str.strip().str.upper()
                            master_df = pd.merge(master_df, report_df, left_on='ASIN', right_on=asin_col,
                                                 how='left', suffixes=suffixes)
                            # Drop duplicate ASIN column if present
                            if asin_col != 'ASIN' and asin_col in master_df.columns:
                                master_df = master_df.drop(columns=[asin_col])
                            print(f"   Merged {asin_col_name} data")
                    return master_df

                df_master = merge_report(df_master, df_weekly, ('', '_weekly'), 'weekly')
                df_master = merge_report(df_master, df_ytd, ('', '_ytd'), 'YTD')
                df_master = merge_report(df_master, df_traffic, ('', '_traffic'), 'traffic')
                df_master = merge_report(df_master, df_inventory, ('', '_inventory'), 'inventory')

                # Set df_weekly to df_master for the rest of the processing
                df_weekly = df_master

                # Save intermediate merged CSV for debugging
                merged_path = os.path.join(FINAL_DATA_DIR,
                                           f"Merged_All_Reports_{year_for_email}_W{week_num_for_email:02d}.csv")
                df_weekly.to_csv(merged_path, index=False)
                print(f"   Saved merged all reports: {os.path.basename(merged_path)}")


                # === HEADER RENAMES (with unique YoY VAR names) ===
                rename_map = {
                    "Ordered Revenue": "WTD Ordered SLS $",
                    "Ordered Revenue - Prior Period (%)": "TY 2wks Ago SLS %",
                    "Ordered Revenue - Same Period Last Year (%)": "YoY VAR SLS",
                    "Ordered Units": "WTD Ordered UNITS",
                    "Ordered Units - Prior Period (%)": "TY 2wks Ago UNITS",
                    "Ordered Units - Same Period Last Year (%)": "YoY VAR UNITS",
                    "Ordered Revenue_ytd": "YTD Ordered SLS $",
                    "Ordered Revenue - Same Period Last Year (%)_ytd": "YoY VAR YTD SLS",
                    "Ordered Units_ytd": "YTD Ordered UNITS",
                    "Ordered Units - Same Period Last Year (%)_ytd": "YoY VAR YTD UNITS",
                    "Featured Offer Page Views": "Glance Views",
                    "Featured Offer Page Views - Prior Period (%)": "TY 2wks Ago Views",
                    "Featured Offer Page Views - Same Period Last Year (%)": "YoY VAR Views",
                    "Open Purchase Order Quantity": "Open PO QTY",
                    "Sellable On Hand Inventory": "On Hand INV $",
                    "Sellable On Hand Units": "On Hand Units",
                    "Sell-Through %": "Sell-Thru %"
                }
                df_weekly = df_weekly.rename(columns=rename_map)
                print("   Applied header renames")

                # === DROP COLUMNS ===
                cols_to_drop = [
                    "Product Title", "Product Title_ytd", "Product Title_traffic", "Product Title_inventory",
                    "Brand", "Brand_ytd", "Brand_traffic", "Brand_inventory",
                    "Shipped Revenue", "Shipped Revenue - Prior Period (%)",
                    "Shipped Revenue - Same Period Last Year (%)",
                    "Shipped COGS", "Shipped COGS - Prior Period (%)",
                    "Shipped COGS - Same Period Last Year (%)",
                    "Shipped Units", "Shipped Units - Prior Period (%)",
                    "Shipped Units - Same Period Last Year (%)",
                    "Customer Returns", "Customer Returns - Prior Period (%)",
                    "Customer Returns - Same Period Last Year (%)",
                    "Customer Returns - Same Period Last Year (%)_ytd",
                    "Shipped Revenue_ytd", "Shipped Revenue - Same Period Last Year (%)_ytd",
                    "Shipped COGS_ytd", "Shipped COGS - Same Period Last Year (%)_ytd",
                    "Shipped Units_ytd", "Shipped Units - Same Period Last Year (%)_ytd",
                    "Customer Returns_ytd", "Customer Returns - Same Period Last Year (%)_ytd",
                    "Sourceable Product OOS %", "Vendor Confirmation %", "Net Received",
                    "Net Received Units", "Receive Fill %", "Overall Vendor Lead Time (days)",
                    "Unfilled Customer Ordered Units",
                    "Aged 90+ Days Sellable Inventory", "Aged 90+ Days Sellable Units",
                    "Unsellable On Hand Inventory", "Unsellable On Hand Units", "Unhealthy Inventory",
                    "Unhealthy Units"
                ]
                existing_drop = [c for c in cols_to_drop if c in df_weekly.columns]
                if existing_drop:
                    df_weekly = df_weekly.drop(columns=existing_drop)
                    print(f"   Dropped {len(existing_drop)} columns")

                    print("   Adding formulaic columns...")

                    # Clean and coerce inputs
                    units = clean_numeric_keep_na(df_weekly['WTD Ordered UNITS'])
                    views = clean_numeric_keep_na(df_weekly['Glance Views'])
                    sls = clean_numeric_keep_na(df_weekly['WTD Ordered SLS $'])
                    open_po_units = clean_numeric_keep_na(df_weekly['Open PO QTY'])
                    on_hand = clean_numeric_keep_na(df_weekly['On Hand Units'])

                    # Formulas (NaN will propagate)
                    df_weekly['Conversion Rate'] = np.where(views == 0, np.nan, (units / views) * 100)
                    df_weekly['ASP $'] = np.where(units == 0, np.nan, sls / units)
                    df_weekly['Wks of INV'] = np.where(units == 0, np.nan, np.floor((on_hand + open_po_units) / units))

                    # === FINAL CLEANING: Converting all numeric columns safely ===
                    print("\n=== FINAL CLEANING: Converting all numeric columns safely ===")

                    # List of all columns that should be numeric
                    numeric_cols = [
                        'WTD Ordered SLS $', 'YTD Ordered SLS $', 'On Hand INV $',
                        'WTD Ordered UNITS', 'YTD Ordered UNITS', 'Open PO QTY', 'On Hand Units',
                        'Glance Views', 'ASP $', 'Wks of INV'
                    ]

                    percent_cols = [
                        'TY 2wks Ago SLS %', 'YoY VAR SLS', 'TY 2wks Ago UNITS', 'YoY VAR UNITS',
                        'YoY VAR YTD SLS', 'YoY VAR YTD UNITS', 'TY 2wks Ago Views', 'YoY VAR Views',
                        'Sell-Thru %', 'Conversion Rate'
                    ]

                    # First: Clean all numeric columns (preserve NaN)
                    for col in numeric_cols:
                        if col in df_weekly.columns:
                            df_weekly[col] = clean_numeric_keep_na(df_weekly[col])

                    # Second: Clean percentage columns, convert to decimal, preserve NaN/blank as NaN
                    for col in percent_cols:
                        if col in df_weekly.columns:
                            cleaned = clean_numeric_keep_na(df_weekly[col])
                            df_weekly[col] = cleaned / 100  # Convert to decimal (e.g., -18.97 → -0.1897)
                            df_weekly[col] = df_weekly[col].replace(0.0, np.nan)  # Optional: exact 0% → blank

                    print("   All numeric and percentage columns cleaned safely")

                    # Drop rows where 'WTD Ordered SLS $', 'YTD Ordered SLS $', and 'On Hand INV $' are all blank (NaN after cleaning)
                    df_weekly = df_weekly.dropna(subset=['WTD Ordered SLS $', 'YTD Ordered SLS $', 'On Hand INV $'],
                                                 how='all')
                    print(
                        f"   Dropped rows where all three sales/inventory columns are blank. New shape: {df_weekly.shape}")

                    # Optional: diagnostic to confirm NaN preservation
                    print("\n=== DIAGNOSTIC: Sample values (NaN should appear where originally blank) ===")
                    sample_cols = ['WTD Ordered SLS $', 'ASP $', 'Conversion Rate', 'Wks of INV',
                                   'Glance Views']
                    if all(c in df_weekly.columns for c in sample_cols):
                        print(df_weekly[sample_cols].head(8).to_string(index=False))
                    else:
                        print("Some diagnostic columns missing")

                    # Reorder columns to the specified order
                    desired_order = [
                        'Model', 'SKUs', 'Description', 'Category', 'ASIN', 'WTD Ordered SLS $', 'TY 2wks Ago SLS %',
                        'YoY VAR SLS', 'WTD Ordered UNITS', 'TY 2wks Ago UNITS', 'YoY VAR UNITS', 'YTD Ordered SLS $',
                        'YoY VAR YTD SLS', 'YTD Ordered UNITS', 'YoY VAR YTD UNITS', 'Open PO QTY',
                        'On Hand INV $', 'On Hand Units', 'Wks of INV', 'Sell-Thru %', 'Glance Views',
                        'TY 2wks Ago Views', 'YoY VAR Views', 'Conversion Rate', 'ASP $'
                    ]

                    # Keep only desired columns (drop others)
                    df_weekly = df_weekly.reindex(columns=desired_order)

                    # === DIAGNOSTIC + CLEAN CONVERSION ===
                    print("\n=== DIAGNOSTIC: Raw values ===")
                    key_cols = [
                        'WTD Ordered SLS $', 'YTD Ordered SLS $', 'On Hand INV $',
                        'TY 2wks Ago SLS %', 'YoY VAR SLS', 'YoY VAR YTD SLS',
                        'WTD Ordered UNITS', 'YTD Ordered UNITS',
                        'TY 2wks Ago UNITS', 'YoY VAR UNITS', 'YoY VAR YTD UNITS',
                        'Glance Views', 'TY 2wks Ago Views', 'YoY VAR Views',
                        'Open PO QTY', 'On Hand Units'
                    ]

                    for col in key_cols:
                        if col in df_weekly.columns:
                            print(f"   {col}: sample={df_weekly[col].head(3).tolist()}")

                    for col in key_cols:
                        if col in df_weekly.columns:
                            df_weekly[col] = clean_numeric_keep_na(df_weekly[col])
                            print(f"   Cleaned {col}: sample={df_weekly[col].head(3).tolist()}")

                    print("\n=== Ready for formatting ===")

                    # Sets $0 weekly values to NaN
                    df_weekly['WTD Ordered SLS $'] = df_weekly['WTD Ordered SLS $'].replace(0, np.nan)

                    # Sort by WTD Ordered SLS $ descending, then by YTD Ordered SLS $ descending
                    df_weekly = df_weekly.sort_values(['WTD Ordered SLS $', 'YTD Ordered SLS $'],
                                                      ascending=[False, False])

                    # === SAVE AS .XLSX WITH FORMATS ===
                    final_report_name = f"Amazon_Weekly_Sales_Report_{year_for_email}_W{week_num_for_email:02d}.xlsx"
                    final_report_path = os.path.join(FINAL_DATA_DIR, final_report_name)

                    df_weekly.to_excel(final_report_path, index=False, engine='openpyxl')

                    from openpyxl import load_workbook
                    from openpyxl.formatting.rule import CellIsRule
                    from openpyxl.styles import Font, Color, Alignment
                    wb = load_workbook(final_report_path)
                    ws = wb.active

                    col_map = {col: idx + 1 for idx, col in enumerate(df_weekly.columns)}

                    # Accounting format
                    accounting_fmt = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
                    accounting_fmt_2dec = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                    for col in ['WTD Ordered SLS $', 'YTD Ordered SLS $', 'On Hand INV $']:
                        if col in col_map:
                            letter = ws.cell(row=1, column=col_map[col]).column_letter
                            for r in range(2, ws.max_row + 1):
                                ws[f"{letter}{r}"].number_format = accounting_fmt

                    if 'ASP $' in col_map:
                        letter = ws.cell(row=1, column=col_map['ASP $']).column_letter
                        for r in range(2, ws.max_row + 1):
                            ws[f"{letter}{r}"].number_format = accounting_fmt_2dec

                    # Percentage format (0.0%)
                    for col in percent_cols:
                        if col in col_map:
                            letter = ws.cell(row=1, column=col_map[col]).column_letter
                            for r in range(2, ws.max_row + 1):
                                try:
                                    ws[f"{letter}{r}"].number_format = '0.0%'
                                except Exception as e:
                                    print(f"Error setting percentage format for column '{col}' row {r}: {e}")

                    # Number format
                    for col in ['WTD Ordered UNITS', 'YTD Ordered UNITS', 'Open PO QTY', 'On Hand Units',
                                'Glance Views']:
                        if col in col_map:
                            letter = ws.cell(row=1, column=col_map[col]).column_letter
                            for r in range(2, ws.max_row + 1):
                                ws[f"{letter}{r}"].number_format = '#,##0'

                    # Bold headers and red negatives for specified columns
                    red_font = Font(color="FF0000")
                    bold_font = Font(bold=True)

                    specified_cols = [
                        'Model',
                        'SKUs',
                        'Description',
                        'TY 2wks Ago SLS %',
                        'YoY VAR SLS',
                        'TY 2wks Ago UNITS',
                        'YoY VAR UNITS',
                        'YTD Ordered SLS $',
                        'YoY VAR YTD SLS',
                        'YoY VAR YTD UNITS',
                        'TY 2wks Ago Views',
                        'YoY VAR Views'
                    ]

                    for col in specified_cols:
                        if col in col_map:
                            col_idx = col_map[col]
                            letter = ws.cell(1, col_idx).column_letter
                            # Bold entire column (header + values)
                            for r in range(1, ws.max_row + 1):
                                ws.cell(r, col_idx).font = bold_font
                            # Conditional red for negatives ( < 0 )
                            ws.conditional_formatting.add(
                                f'{letter}2:{letter}{ws.max_row}',
                                CellIsRule(operator='lessThan', formula=['0'], font=red_font)
                            )

                    # Auto-fit columns A-D
                    for col in range(1, 5):  # A=1, B=2, C=3, D=4
                        max_length = 0
                        column_letter = ws.cell(1, col).column_letter
                        for row in range(1, ws.max_row + 1):
                            cell = ws.cell(row, col)
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws.column_dimensions[column_letter].width = max_length * 1.2  # Buffer for readability

                    # Set E+ to width 11 and wrap header text
                    for col in range(5, ws.max_column + 1):
                        column_letter = ws.cell(1, col).column_letter
                        ws.column_dimensions[column_letter].width = 11
                        # Wrap header text
                        ws.cell(1, col).alignment = Alignment(wrap_text=True, vertical='center',
                                                              horizontal='center')

                    wb.save(final_report_path)
                    print(f"   ✅ Saved {final_report_name} with correct Accounting and Percentage formats")
                    print(f"   Final shape: {df_weekly.shape}")

                    #===================================================
                    # Generate Pie Charts
                    #===================================================

                    # Prepare consistent color map for categories (unchanged)
                    unique_categories = sorted(df_weekly['Category'].unique())
                    cat_color_map = {cat: plt.cm.tab10(i / len(unique_categories)) for i, cat in
                                     enumerate(unique_categories)}

                    # Specify colors for top categories (unchanged)
                    cat_color_map['Lithion'] = '#06b553'  # Zithion Green
                    cat_color_map['Headlamp'] = '#1a6079'  # Ocean
                    cat_color_map['Flashlight'] = '#8d6f51'  # Desert
                    cat_color_map['Work Light'] = '#e62e2c'  # Crimson
                    cat_color_map['Lantern'] = '#998715'  # Warm Light
                    cat_color_map['Folding'] = '#027360'  # Teal
                    cat_color_map['Eyewear'] = '#59144D'  # Purple
                    cat_color_map['Other'] = '#808080'  # Medium Gray


                    # Define common variables (unchanged)
                    label_distance = 1.2
                    line_height = 0.05  # Reduced line height for tighter spacing

                    # Generate YTD pie chart (Revenue) - use distinct var
                    ytd_rev_sum = df_weekly.groupby('Category')['YTD Ordered SLS $'].sum().reset_index()
                    ytd_rev_sum = ytd_rev_sum[ytd_rev_sum['YTD Ordered SLS $'] > 0].sort_values('YTD Ordered SLS $',
                                                                                                ascending=False)
                    fig, ax = plt.subplots(figsize=(13, 13))  # Square figure for better spacing
                    fig.subplots_adjust(left=0.05, right=0.95, bottom=0.1, top=0.95)
                    fig.patch.set_edgecolor('black')
                    fig.patch.set_linewidth(2)
                    top_n = 5
                    if len(ytd_rev_sum) > top_n:
                        top_categories = ytd_rev_sum.iloc[:top_n]['Category'].tolist()
                        top_values = ytd_rev_sum.iloc[:top_n]['YTD Ordered SLS $'].tolist()
                        other_value = ytd_rev_sum.iloc[top_n:]['YTD Ordered SLS $'].sum()
                        top_categories.append('Other')
                        top_values.append(other_value)
                    else:
                        top_categories = ytd_rev_sum['Category'].tolist()
                        top_values = ytd_rev_sum['YTD Ordered SLS $'].tolist()
                    # Explode: base 0.1, increase for smaller slices (last 3 get more)
                    explode = [0.05] * len(top_values)
                    if len(explode) > 2:
                        explode[-4:] = [0.075, 0.075, 0.075, 0.075]  # Slightly more for smallest 3
                    total = sum(top_values)
                    pcts = [v / total * 100 for v in top_values]
                    pie_colors = [cat_color_map[cat] for cat in top_categories]
                    wedges, _ = ax.pie(top_values, explode=explode, colors=pie_colors,
                                       shadow=True, startangle=90,
                                       wedgeprops=dict(width=0.5, edgecolor='white', linewidth=1))
                    # Manually add labels with bold category
                    for i, wedge in enumerate(wedges):
                        ang = (wedge.theta2 - wedge.theta1) / 2. + wedge.theta1
                        y = np.sin(np.deg2rad(ang))
                        x = np.cos(np.deg2rad(ang))
                        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
                        x_label = label_distance * x
                        y_label = label_distance * y
                        cat = top_categories[i]
                        v = top_values[i]
                        p = pcts[i]
                        # Category (bold)
                        ax.text(x_label, y_label + line_height, cat, ha=horizontalalignment, va='center',
                                fontsize=12, fontweight='bold')
                        # Value
                        ax.text(x_label, y_label, f'${int(v):,}', ha=horizontalalignment, va='center', fontsize=12)
                        # Percentage
                        ax.text(x_label, y_label - line_height, f'{p:.1f}%', ha=horizontalalignment, va='center',
                                fontsize=12)
                    # Move title to bottom
                    ax.text(0.5, -0.1, 'YTD Revenue by Category', ha='center', fontsize=24, fontweight='bold',
                            transform=ax.transAxes)
                    ytd_rev_pie_img_path = os.path.join(os.environ.get('SALES_DATA_DIR', './data'), 'PNG files\YTD Sales by Category Revenue.png'
                    plt.savefig(ytd_rev_pie_img_path, format='png')

                    # Generate WTD pie chart (Revenue) - use distinct var
                    wtd_rev_sum = df_weekly.groupby('Category')['WTD Ordered SLS $'].sum().reset_index()
                    wtd_rev_sum = wtd_rev_sum[wtd_rev_sum['WTD Ordered SLS $'] > 0].sort_values('WTD Ordered SLS $',
                                                                                                ascending=False)
                    wtd_fig, wtd_ax = plt.subplots(figsize=(13, 13))  # Square figure for better spacing
                    wtd_fig.subplots_adjust(left=0.05, right=0.95, bottom=0.1, top=0.95)
                    wtd_fig.patch.set_edgecolor('black')
                    wtd_fig.patch.set_linewidth(2)
                    if len(wtd_rev_sum) > top_n:
                        wtd_top_categories = wtd_rev_sum.iloc[:top_n]['Category'].tolist()
                        wtd_top_values = wtd_rev_sum.iloc[:top_n]['WTD Ordered SLS $'].tolist()
                        wtd_other_value = wtd_rev_sum.iloc[top_n:]['WTD Ordered SLS $'].sum()
                        wtd_top_categories.append('Other')
                        wtd_top_values.append(wtd_other_value)
                    else:
                        wtd_top_categories = wtd_rev_sum['Category'].tolist()
                        wtd_top_values = wtd_rev_sum['WTD Ordered SLS $'].tolist()

                    # Explode: base 0.1, increase for smaller slices (last 3 get more)
                    wtd_explode = [0.05] * len(wtd_top_values)
                    if len(wtd_explode) > 2:
                        wtd_explode[-4:] = [0.075, 0.075, 0.075, 0.075]  # Slightly more for smallest 3
                    wtd_total = sum(wtd_top_values)
                    wtd_pcts = [v / wtd_total * 100 for v in wtd_top_values]
                    wtd_pie_colors = [cat_color_map[cat] for cat in wtd_top_categories]
                    wtd_wedges, _ = wtd_ax.pie(wtd_top_values, explode=wtd_explode, colors=wtd_pie_colors,
                                               shadow=True, startangle=90,
                                               wedgeprops=dict(width=0.5, edgecolor='white', linewidth=1))
                    # Manually add labels with bold category
                    for i, wedge in enumerate(wtd_wedges):
                        ang = (wedge.theta2 - wedge.theta1) / 2. + wedge.theta1
                        y = np.sin(np.deg2rad(ang))
                        x = np.cos(np.deg2rad(ang))
                        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
                        x_label = label_distance * x
                        y_label = label_distance * y
                        cat = wtd_top_categories[i]
                        v = wtd_top_values[i]
                        p = wtd_pcts[i]
                        # Category (bold)
                        wtd_ax.text(x_label, y_label + line_height, cat, ha=horizontalalignment, va='center',
                                    fontsize=12, fontweight='bold')
                        # Value
                        wtd_ax.text(x_label, y_label, f'${int(v):,}', ha=horizontalalignment, va='center',
                                    fontsize=12)
                        # Percentage
                        wtd_ax.text(x_label, y_label - line_height, f'{p:.1f}%', ha=horizontalalignment,
                                    va='center', fontsize=12)
                    # Move title to bottom
                    wtd_ax.text(0.5, -0.1, 'Weekly Revenue by Category', ha='center', fontsize=24,
                                fontweight='bold',
                                transform=wtd_ax.transAxes)
                    week_rev_pie_img_path = os.path.join(os.environ.get('SALES_DATA_DIR', './data'), 'PNG files\Weekly Sales by Category Revenue.png'
                    plt.savefig(week_rev_pie_img_path, format='png')

                    # Generate YTD pie chart (Units) - use distinct var
                    ytd_units_sum = df_weekly.groupby('Category')['YTD Ordered UNITS'].sum().reset_index()
                    ytd_units_sum = ytd_units_sum[ytd_units_sum['YTD Ordered UNITS'] > 0].sort_values(
                        'YTD Ordered UNITS', ascending=False)
                    fig, ax = plt.subplots(figsize=(13, 13))  # Square figure for better spacing
                    fig.subplots_adjust(left=0.05, right=0.95, bottom=0.1, top=0.95)
                    fig.patch.set_edgecolor('black')
                    fig.patch.set_linewidth(2)
                    top_n = 5
                    if len(ytd_units_sum) > top_n:
                        top_categories = ytd_units_sum.iloc[:top_n]['Category'].tolist()
                        top_values = ytd_units_sum.iloc[:top_n]['YTD Ordered UNITS'].tolist()
                        other_value = ytd_units_sum.iloc[top_n:]['YTD Ordered UNITS'].sum()
                        top_categories.append('Other')
                        top_values.append(other_value)
                    else:
                        top_categories = ytd_units_sum['Category'].tolist()
                        top_values = ytd_units_sum['YTD Ordered UNITS'].tolist()
                    # Explode: base 0.1, increase for smaller slices (last 3 get more)
                    explode = [0.05] * len(top_values)
                    if len(explode) > 2:
                        explode[-4:] = [0.075, 0.075, 0.075, 0.075]  # Slightly more for smallest 3
                    total = sum(top_values)
                    pcts = [v / total * 100 for v in top_values]
                    pie_colors = [cat_color_map[cat] for cat in top_categories]
                    wedges, _ = ax.pie(top_values, explode=explode, colors=pie_colors,
                                       shadow=True, startangle=90,
                                       wedgeprops=dict(width=0.5, edgecolor='white', linewidth=1))
                    # Manually add labels with bold category
                    for i, wedge in enumerate(wedges):
                        ang = (wedge.theta2 - wedge.theta1) / 2. + wedge.theta1
                        y = np.sin(np.deg2rad(ang))
                        x = np.cos(np.deg2rad(ang))
                        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
                        x_label = label_distance * x
                        y_label = label_distance * y
                        cat = top_categories[i]
                        v = top_values[i]
                        p = pcts[i]
                        # Category (bold)
                        ax.text(x_label, y_label + line_height, cat, ha=horizontalalignment, va='center',
                                fontsize=12, fontweight='bold')
                        # Value (no $ for units)
                        ax.text(x_label, y_label, f'{int(v):,}', ha=horizontalalignment, va='center', fontsize=12)
                        # Percentage
                        ax.text(x_label, y_label - line_height, f'{p:.1f}%', ha=horizontalalignment, va='center',
                                fontsize=12)
                    # Move title to bottom
                    ax.text(0.5, -0.1, 'YTD Units by Category', ha='center', fontsize=24, fontweight='bold',
                            transform=ax.transAxes)
                    ytd_units_pie_img_path = os.path.join(os.environ.get('SALES_DATA_DIR', './data'), 'PNG files\YTD Sales by Category Units.png'
                    plt.savefig(ytd_units_pie_img_path, format='png')

                    # Generate WTD pie chart (Units) - use distinct var
                    wtd_units_sum = df_weekly.groupby('Category')['WTD Ordered UNITS'].sum().reset_index()
                    wtd_units_sum = wtd_units_sum[wtd_units_sum['WTD Ordered UNITS'] > 0].sort_values(
                        'WTD Ordered UNITS', ascending=False)
                    wtd_fig, wtd_ax = plt.subplots(figsize=(13, 13))  # Square figure for better spacing
                    wtd_fig.subplots_adjust(left=0.05, right=0.95, bottom=0.1, top=0.95)
                    wtd_fig.patch.set_edgecolor('black')
                    wtd_fig.patch.set_linewidth(2)
                    if len(wtd_units_sum) > top_n:
                        wtd_top_categories = wtd_units_sum.iloc[:top_n]['Category'].tolist()
                        wtd_top_values = wtd_units_sum.iloc[:top_n]['WTD Ordered UNITS'].tolist()
                        wtd_other_value = wtd_units_sum.iloc[top_n:]['WTD Ordered UNITS'].sum()
                        wtd_top_categories.append('Other')
                        wtd_top_values.append(wtd_other_value)
                    else:
                        wtd_top_categories = wtd_units_sum['Category'].tolist()
                        wtd_top_values = wtd_units_sum['WTD Ordered UNITS'].tolist()

                    # Explode: base 0.1, increase for smaller slices (last 3 get more)
                    wtd_explode = [0.05] * len(wtd_top_values)
                    if len(wtd_explode) > 2:
                        wtd_explode[-4:] = [0.075, 0.075, 0.075, 0.075]  # Slightly more for smallest 3
                    wtd_total = sum(wtd_top_values)
                    wtd_pcts = [v / wtd_total * 100 for v in wtd_top_values]
                    wtd_pie_colors = [cat_color_map[cat] for cat in wtd_top_categories]
                    wtd_wedges, _ = wtd_ax.pie(wtd_top_values, explode=wtd_explode, colors=wtd_pie_colors,
                                               shadow=True, startangle=90,
                                               wedgeprops=dict(width=0.5, edgecolor='white', linewidth=1))
                    # Manually add labels with bold category
                    for i, wedge in enumerate(wtd_wedges):
                        ang = (wedge.theta2 - wedge.theta1) / 2. + wedge.theta1
                        y = np.sin(np.deg2rad(ang))
                        x = np.cos(np.deg2rad(ang))
                        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
                        x_label = label_distance * x
                        y_label = label_distance * y
                        cat = wtd_top_categories[i]
                        v = wtd_top_values[i]
                        p = wtd_pcts[i]
                        # Category (bold)
                        wtd_ax.text(x_label, y_label + line_height, cat, ha=horizontalalignment, va='center',
                                    fontsize=12, fontweight='bold')
                        # Value (no $ for units)
                        wtd_ax.text(x_label, y_label, f'{int(v):,}', ha=horizontalalignment, va='center',
                                    fontsize=12)
                        # Percentage
                        wtd_ax.text(x_label, y_label - line_height, f'{p:.1f}%', ha=horizontalalignment,
                                    va='center', fontsize=12)
                    # Move title to bottom
                    wtd_ax.text(0.5, -0.1, 'Weekly Units by Category', ha='center', fontsize=24,
                                fontweight='bold',
                                transform=wtd_ax.transAxes)
                    week_units_pie_img_path = os.path.join(os.environ.get('SALES_DATA_DIR', './data'), 'PNG files\Weekly Sales by Category Units.png'
                    plt.savefig(week_units_pie_img_path, format='png')

                    # Resize all four PNG files to 65% of their original size and resave them (reduced from 90% for better email fit)
                    image_paths = [
                        ytd_rev_pie_img_path,
                        week_rev_pie_img_path,
                        ytd_units_pie_img_path,
                        week_units_pie_img_path
                    ]

                    for path in image_paths:
                        img = mpimg.imread(path)
                        original_shape = img.shape
                        resized_img = zoom(img, (0.65, 0.65, 1) if len(img.shape) == 3 else (0.65, 0.65))
                        resized_img = np.clip(resized_img, 0.0, 1.0)
                        mpimg.imsave(path, resized_img)
                        new_shape = resized_img.shape
                        print(f"Resized {path} from {original_shape} to {new_shape}")

                    # Close all figures to free up memory
                    plt.close('all')

                    # ==============================================================
                    #    UPDATE CATEGORY WEEKLY EXCEL FILES
                    # ==============================================================
                    revenue_excel_path = os.path.join(os.environ.get("SALES_DATA_DIR", "./data"), "Category Weekly Revenue Data.xlsx"
                    units_excel_path = os.path.join(os.environ.get("SALES_DATA_DIR", "./data"), "Category Weekly Units Data.xlsx"

                    # === FIX: Use the most recent completed week's number ===
                    # successful_weeks[0] is the most recent (since appended starting from most recent)
                    most_recent_sunday = successful_weeks[0][0]
                    most_recent_week_info = get_business_week_info(datetime.strptime(most_recent_sunday, "%Y-%m-%d"))
                    week_num = most_recent_week_info["week_number"]

                    print(f"Updating category weekly files for Week {week_num}")

                    # For revenue
                    df_rev = pd.read_excel(revenue_excel_path)

                    # Clean 'Week' column to ensure numeric
                    df_rev['Week'] = pd.to_numeric(df_rev['Week'], errors='coerce')

                    if week_num in df_rev['Week'].values:
                        week_row = df_rev[df_rev['Week'] == week_num].index[0]
                    else:
                        week_row = len(df_rev)
                        df_rev.loc[week_row, 'Week'] = week_num

                    print(f"Revenue DataFrame before update:\n{df_rev.to_string(index=False)}")

                    for _, row in wtd_rev_sum.iterrows():
                        cat = row['Category']
                        val = row['WTD Ordered SLS $']
                        if pd.isna(val):
                            val = 0
                        if cat in df_rev.columns:
                            df_rev.loc[week_row, cat] = val
                            print(f"Set {cat} for week {week_num} to {val}")

                    # Fill any remaining NaN in the row with 0
                    df_rev.loc[week_row] = df_rev.loc[week_row].fillna(0)

                    print(f"Revenue DataFrame after update:\n{df_rev.to_string(index=False)}")

                    df_rev.to_excel(revenue_excel_path, index=False)
                    print(f"   Updated {os.path.basename(revenue_excel_path)} with week {week_num}")

                    # For units
                    df_units = pd.read_excel(units_excel_path)

                    # Clean 'Week' column to ensure numeric
                    df_units['Week'] = pd.to_numeric(df_units['Week'], errors='coerce')

                    if week_num in df_units['Week'].values:
                        week_row = df_units[df_units['Week'] == week_num].index[0]
                    else:
                        week_row = len(df_units)
                        df_units.loc[week_row, 'Week'] = week_num

                    print(f"Units DataFrame before update:\n{df_units.to_string(index=False)}")

                    for _, row in wtd_units_sum.iterrows():
                        cat = row['Category']
                        val = row['WTD Ordered UNITS']
                        if pd.isna(val):
                            val = 0
                        if cat in df_units.columns:
                            df_units.loc[week_row, cat] = val
                            print(f"Set {cat} for week {week_num} to {val}")

                    # Fill any remaining NaN in the row with 0
                    df_units.loc[week_row] = df_units.loc[week_row].fillna(0)

                    print(f"Units DataFrame after update:\n{df_units.to_string(index=False)}")

                    df_units.to_excel(units_excel_path, index=False)
                    print(f"   Updated {os.path.basename(units_excel_path)} with week {week_num}")

                    ###############################################################
                    # GENERATE CATAGORICAL LINE GRAPHS
                    ###############################################################

                    # Now, generate line graphs
                    # For revenue
                    fig, ax = plt.subplots(figsize=(12, 7.5))
                    fig.patch.set_edgecolor('black')
                    fig.patch.set_linewidth(2)
                    df_rev = pd.read_excel(revenue_excel_path)

                    # Clean data: convert numeric columns
                    for col in df_rev.columns[1:]:  # Skip 'Week'
                        df_rev[col] = pd.to_numeric(df_rev[col], errors='coerce')
                        df_rev[col] = df_rev[col].fillna(0)  # Fill NaN with 0 for summation and plotting

                    # Filter to only include weeks up to the current week_num
                    df_rev = df_rev[df_rev['Week'] <= week_num].copy()

                    df_rev.set_index('Week', inplace=True)

                    top_cats_rev = wtd_rev_sum.sort_values('WTD Ordered SLS $', ascending=False)['Category'][
                                   :7].tolist()
                    other_cats_rev = [c for c in df_rev.columns if c not in top_cats_rev]

                    # Create a list to hold handles and labels for custom legend
                    handles = []
                    labels = []

                    for cat in top_cats_rev:
                        latest_val = df_rev[cat].iloc[-1] if not df_rev.empty else 0
                        label = f"{cat} (${int(latest_val):,})"
                        line, = ax.plot(df_rev.index, df_rev[cat], label=label, color=cat_color_map.get(cat, 'gray'),
                                        linewidth=3.5)
                        handles.append(line)
                        labels.append(label)

                    # Other sum
                    df_rev['Other'] = df_rev[other_cats_rev].sum(axis=1)
                    latest_other = df_rev['Other'].iloc[-1] if not df_rev.empty else 0
                    other_label = f"Other (${int(latest_other):,})"
                    other_line, = ax.plot(df_rev.index, df_rev['Other'], label=other_label,
                                          color=cat_color_map.get('Other', 'gray'),
                                          linewidth=3.5)
                    handles.append(other_line)
                    labels.append(other_label)

                    ax.set_title('Weekly Revenue by Category', fontsize=14, fontweight='bold')
                    ax.set_xlabel('Week Number')
                    ax.set_ylabel('Revenue $')
                    ax.yaxis.set_major_formatter(
                        plt.FuncFormatter(lambda x, _: f'${int(x):,}' if not math.isnan(x) else ''))
                    ax.grid(True, linestyle='--', alpha=0.7)
                    ax.set_xticks(df_rev.index)

                    # Sort legend by latest values descending
                    # Create a list of tuples: (latest_val, handle, label)
                    sorted_items = []
                    for h, l in zip(handles, labels):
                        # Extract value from label, e.g., "Category ($123,456)" → 123456
                        val_str = l.split('($')[1].rstrip(')') if '($' in l else '0'
                        val = int(val_str.replace(',', '')) if val_str else 0
                        sorted_items.append((val, h, l))

                    # Sort descending by value
                    sorted_items.sort(key=lambda x: x[0], reverse=True)

                    # Extract sorted handles and labels
                    sorted_handles = [item[1] for item in sorted_items]
                    sorted_labels = [item[2] for item in sorted_items]

                    # Add sorted legend
                    ax.legend(sorted_handles, sorted_labels, loc='upper left', frameon=True)

                    category_rev_line_path = os.path.join(os.environ.get('SALES_DATA_DIR', './data'), 'PNG Files\Category Weekly Revenue Line.png'
                    plt.savefig(category_rev_line_path, bbox_inches='tight')
                    plt.close(fig)

                    # For cumulative revenue
                    fig, ax = plt.subplots(figsize=(12, 7.5))
                    fig.patch.set_edgecolor('black')
                    fig.patch.set_linewidth(2)

                    df_rev_cum = df_rev.cumsum()

                    # Reuse the same top categories for consistency
                    handles = []
                    labels = []

                    for cat in top_cats_rev:
                        latest_val = df_rev_cum[cat].iloc[-1] if not df_rev_cum.empty else 0
                        label = f"{cat} (${int(latest_val):,})"
                        line, = ax.plot(df_rev_cum.index, df_rev_cum[cat], label=label,
                                        color=cat_color_map.get(cat, 'gray'),
                                        linewidth=3.5)
                        handles.append(line)
                        labels.append(label)

                    # Other sum (cumulative)
                    df_rev_cum['Other'] = df_rev_cum['Other']  # Already cumsum'd from df_rev
                    latest_other = df_rev_cum['Other'].iloc[-1] if not df_rev_cum.empty else 0
                    other_label = f"Other (${int(latest_other):,})"
                    other_line, = ax.plot(df_rev_cum.index, df_rev_cum['Other'], label=other_label,
                                          color=cat_color_map.get('Other', 'gray'),
                                          linewidth=3.5)
                    handles.append(other_line)
                    labels.append(other_label)

                    ax.set_title('Cumulative Revenue by Category', fontsize=14, fontweight='bold')
                    ax.set_xlabel('Week Number')
                    ax.set_ylabel('Revenue $')
                    ax.yaxis.set_major_formatter(
                        plt.FuncFormatter(lambda x, _: f'${int(x):,}' if not math.isnan(x) else ''))
                    ax.grid(True, linestyle='--', alpha=0.7)
                    ax.set_xticks(df_rev_cum.index)

                    # Sort legend by latest values descending
                    sorted_items = []
                    for h, l in zip(handles, labels):
                        val_str = l.split('($')[1].rstrip(')') if '($' in l else '0'
                        val = int(val_str.replace(',', '')) if val_str else 0
                        sorted_items.append((val, h, l))

                    sorted_items.sort(key=lambda x: x[0], reverse=True)

                    sorted_handles = [item[1] for item in sorted_items]
                    sorted_labels = [item[2] for item in sorted_items]

                    ax.legend(sorted_handles, sorted_labels, loc='upper left', frameon=True)

                    category_cum_rev_line_path = os.path.join(os.environ.get('SALES_DATA_DIR', './data'), 'PNG Files\Category Cumulative Revenue Line.png'
                    plt.savefig(category_cum_rev_line_path, bbox_inches='tight')
                    plt.close(fig)

                    # For units
                    fig, ax = plt.subplots(figsize=(12, 7.5))
                    fig.patch.set_edgecolor('black')
                    fig.patch.set_linewidth(2)
                    df_units = pd.read_excel(units_excel_path)

                    # Clean data: convert numeric columns
                    for col in df_units.columns[1:]:  # Skip 'Week'
                        df_units[col] = pd.to_numeric(df_units[col], errors='coerce')
                        df_units[col] = df_units[col].fillna(0)  # Fill NaN with 0 for summation and plotting

                    # Filter to only include weeks up to the current week_num
                    df_units = df_units[df_units['Week'] <= week_num].copy()

                    df_units.set_index('Week', inplace=True)

                    top_cats_units = wtd_units_sum.sort_values('WTD Ordered UNITS', ascending=False)['Category'][
                                     :7].tolist()
                    other_cats_units = [c for c in df_units.columns if c not in top_cats_units]

                    # Create a list to hold handles and labels for custom legend
                    handles = []
                    labels = []

                    for cat in top_cats_units:
                        latest_val = df_units[cat].iloc[-1] if not df_units.empty else 0
                        label = f"{cat} ({int(latest_val):,})"
                        line, = ax.plot(df_units.index, df_units[cat], label=label,
                                        color=cat_color_map.get(cat, 'gray'),
                                        linewidth=3.5)
                        handles.append(line)
                        labels.append(label)

                    # Other sum
                    df_units['Other'] = df_units[other_cats_units].sum(axis=1)
                    latest_other = df_units['Other'].iloc[-1] if not df_units.empty else 0
                    other_label = f"Other ({int(latest_other):,})"
                    other_line, = ax.plot(df_units.index, df_units['Other'], label=other_label,
                                          color=cat_color_map.get('Other', 'gray'),
                                          linewidth=3.5)
                    handles.append(other_line)
                    labels.append(other_label)

                    ax.set_title('Weekly Units by Category', fontsize=14, fontweight='bold')
                    ax.set_xlabel('Week Number')
                    ax.set_ylabel('Units')
                    ax.yaxis.set_major_formatter(
                        plt.FuncFormatter(lambda x, _: f'{int(x):,}' if not math.isnan(x) else ''))
                    ax.grid(True, linestyle='--', alpha=0.7)
                    ax.set_xticks(df_units.index)

                    # Sort legend by latest values descending
                    # Create a list of tuples: (latest_val, handle, label)
                    sorted_items = []
                    for h, l in zip(handles, labels):
                        # Extract value from label, e.g., "Category (123,456)" → 123456
                        val_str = l.split('(')[1].rstrip(')') if '(' in l else '0'
                        val = int(val_str.replace(',', '')) if val_str else 0
                        sorted_items.append((val, h, l))

                    # Sort descending by value
                    sorted_items.sort(key=lambda x: x[0], reverse=True)

                    # Extract sorted handles and labels
                    sorted_handles = [item[1] for item in sorted_items]
                    sorted_labels = [item[2] for item in sorted_items]

                    # Add sorted legend
                    ax.legend(sorted_handles, sorted_labels, loc='upper left', frameon=True)

                    category_units_line_path = os.path.join(os.environ.get('SALES_DATA_DIR', './data'), 'PNG Files\Category Weekly Units Line.png'
                    plt.savefig(category_units_line_path, bbox_inches='tight')
                    plt.close(fig)

                    # For cumulative units
                    fig, ax = plt.subplots(figsize=(12, 7.5))
                    fig.patch.set_edgecolor('black')
                    fig.patch.set_linewidth(2)

                    df_units_cum = df_units.cumsum()

                    # Reuse the same top categories for consistency
                    handles = []
                    labels = []

                    for cat in top_cats_units:
                        latest_val = df_units_cum[cat].iloc[-1] if not df_units_cum.empty else 0
                        label = f"{cat} ({int(latest_val):,})"
                        line, = ax.plot(df_units_cum.index, df_units_cum[cat], label=label,
                                        color=cat_color_map.get(cat, 'gray'),
                                        linewidth=3.5)
                        handles.append(line)
                        labels.append(label)

                    # Other sum (cumulative)
                    df_units_cum['Other'] = df_units_cum['Other']  # Already cumsum'd from df_units
                    latest_other = df_units_cum['Other'].iloc[-1] if not df_units_cum.empty else 0
                    other_label = f"Other ({int(latest_other):,})"
                    other_line, = ax.plot(df_units_cum.index, df_units_cum['Other'], label=other_label,
                                          color=cat_color_map.get('Other', 'gray'),
                                          linewidth=3.5)
                    handles.append(other_line)
                    labels.append(other_label)

                    ax.set_title('Cumulative Units by Category', fontsize=14, fontweight='bold')
                    ax.set_xlabel('Week Number')
                    ax.set_ylabel('Units')
                    ax.yaxis.set_major_formatter(
                        plt.FuncFormatter(lambda x, _: f'{int(x):,}' if not math.isnan(x) else ''))
                    ax.grid(True, linestyle='--', alpha=0.7)
                    ax.set_xticks(df_units_cum.index)

                    # Sort legend by latest values descending
                    sorted_items = []
                    for h, l in zip(handles, labels):
                        val_str = l.split('(')[1].rstrip(')') if '(' in l else '0'
                        val = int(val_str.replace(',', '')) if val_str else 0
                        sorted_items.append((val, h, l))

                    sorted_items.sort(key=lambda x: x[0], reverse=True)

                    sorted_handles = [item[1] for item in sorted_items]
                    sorted_labels = [item[2] for item in sorted_items]

                    ax.legend(sorted_handles, sorted_labels, loc='upper left', frameon=True)

                    category_cum_units_line_path = os.path.join(os.environ.get('SALES_DATA_DIR', './data'), 'PNG Files\Category Cumulative Units Line.png'
                    plt.savefig(category_cum_units_line_path, bbox_inches='tight')
                    plt.close(fig)

                    # Resize them to 75%
                    for path in [category_rev_line_path, category_cum_rev_line_path, category_units_line_path,
                                 category_cum_units_line_path]:
                        img = mpimg.imread(path)
                        resized_img = zoom(img, (0.75, 0.75, 1) if len(img.shape) == 3 else (0.75, 0.75))
                        resized_img = np.clip(resized_img, 0.0, 1.0)
                        mpimg.imsave(path, resized_img)

                    # # ===================================================
                    # # Integrate NetSuite Download and Processing (Commented out until COGS data in NetSuite is fixed)
                    # # ===================================================
                    # print("\n" + "=" * 90)
                    # print("          STARTING NETSUITE LOGIN AND DOWNLOAD")
                    # print("=" * 90)
                    #
                    # # Use a new driver for NetSuite to avoid interfering with Amazon session
                    # driver_netsuite = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
                    #
                    # # NetSuite login
                    # driver_netsuite.get("https://4423908.app.netsuite.com")  # Adjust if needed
                    # log("Starting NetSuite login...")
                    # netsuite_login(driver_netsuite, log)
                    # log("NetSuite login successful!")
                    #
                    # # Navigate to the search results page
                    # netsuite_url = "https://4423908.app.netsuite.com/app/common/search/searchresults.nl?searchid=10957"
                    # driver_netsuite.get(netsuite_url)
                    # time.sleep(10)  # Wait for page to load
                    #
                    # # Click the Export CSV button
                    # try:
                    #     export_button = WebDriverWait(driver_netsuite, 30).until(
                    #         EC.element_to_be_clickable((By.CSS_SELECTOR, ".uir-list-icon-button.uir-list-export-csv"))
                    #     )
                    #     driver_netsuite.execute_script("arguments[0].scrollIntoView({block: 'center'});", export_button)
                    #     driver_netsuite.execute_script("arguments[0].click();", export_button)
                    #     print("   Clicked Export - CSV button")
                    # except Exception as e:
                    #     print(f"   Failed to click Export button: {e}")
                    #     driver_netsuite.save_screenshot("ERROR_NETSUITE_EXPORT.png")
                    #
                    # # Wait for the download
                    # downloaded_netsuite = None
                    # start_time_ns = time.time()  # Separate start_time for NetSuite wait
                    # while time.time() - start_time_ns < 120:
                    #     netsuite_files = sorted(Path(DOWNLOAD_DIR).glob("HardistyAmazonRecentItems*.csv"), key=os.path.getmtime, reverse=True)
                    #     if netsuite_files:
                    #         latest_file = netsuite_files[0]
                    #         if "crdownload" not in latest_file.name:
                    #             downloaded_netsuite = str(latest_file)
                    #             break
                    #     time.sleep(1)
                    #
                    # if downloaded_netsuite:
                    #     print(f"   NetSuite file downloaded: {os.path.basename(downloaded_netsuite)}")
                    # else:
                    #     print("   NetSuite file not detected — check manually")
                    #
                    # # Close NetSuite driver
                    # driver_netsuite.quit()
                    #
                    # # Now process the NetSuite file to add COGS, Margin, etc. to Amazon report
                    # if downloaded_netsuite:
                    #     print("\n" + "=" * 90)
                    #     print("          PROCESSING NETSUITE DATA AND UPDATING AMAZON REPORT")
                    #     print("=" * 90)
                    #
                    #     df_netsuite = pd.read_csv(downloaded_netsuite)
                    #
                    #     # Create dictionary: Item -> COGS (convert COGS to numeric)
                    #     df_netsuite['COGS'] = pd.to_numeric(df_netsuite['COGS'], errors='coerce')
                    #     cogs_dict = dict(zip(df_netsuite['Item'].astype(str).str.strip(), df_netsuite['COGS']))
                    #
                    #     # Load the Amazon XLSX
                    #     amazon_path = os.path.join(FINAL_DATA_DIR, f"Amazon_Weekly_Sales_Report_{year_for_email}_W{week_num_for_email:02d}.xlsx")
                    #     df_amazon = pd.read_excel(amazon_path)
                    #
                    #     # Function to get average COGS for a SKUs string
                    #     def get_average_cogs(skus_str):
                    #         if pd.isna(skus_str):
                    #             return np.nan
                    #         skus = [s.strip() for s in str(skus_str).split(',')]  # Assume commas separate multiple SKUs
                    #         valid_cogs = [cogs_dict.get(s) for s in skus if s in cogs_dict and not pd.isna(cogs_dict[s])]
                    #         if valid_cogs:
                    #             return np.mean(valid_cogs)
                    #         return np.nan
                    #
                    #     # Add COGS column
                    #     df_amazon['COGS'] = df_amazon['SKUs'].apply(get_average_cogs)
                    #
                    #     # Ensure ASP $ is numeric
                    #     df_amazon['ASP $'] = pd.to_numeric(df_amazon['ASP $'], errors='coerce')
                    #
                    #     # Add Margin ($) = ASP $ - COGS
                    #     df_amazon['Margin ($)'] = df_amazon['ASP $'] - df_amazon['COGS']
                    #
                    #     # Add Margin (%) = Margin ($) / ASP $ (handle division by zero)
                    #     df_amazon['Margin (%)'] = np.where(df_amazon['ASP $'] != 0, df_amazon['Margin ($)'] / df_amazon['ASP $'], np.nan)
                    #
                    #     # Save back to the XLSX
                    #     df_amazon.to_excel(amazon_path, index=False)
                    #
                    #     print(f"Updated Amazon_Weekly_Sales_Report_{year_for_email}_W{week_num_for_email:02d}.xlsx with COGS, Margin ($), and Margin (%) columns.")
                    #
                    #     # Rename and move the NetSuite file
                    #     cogs_file_name = f"W{week_num_for_email:02d}_Amazon_COGS_Data.csv"
                    #     cogs_dest = os.path.join(FINAL_DATA_DIR, cogs_file_name)
                    #     os.replace(downloaded_netsuite, cogs_dest)
                    #     print(f"   Renamed and moved NetSuite file to {cogs_file_name} in Sales Data folder")

                    # ==============================================================
                    #    ADDING TOTAL ROW (after NetSuite update)
                    # ==============================================================

                    # Reload updated df_weekly from XLSX
                    df_weekly = pd.read_excel(final_report_path)

                    # Compute total row
                    total_row = pd.Series(index=df_weekly.columns, dtype=object)
                    total_row['Model'] = 'Total'

                    # avg_cols = ['COGS', 'Margin ($)', 'Margin (%)']
                    #
                    # # Average other numeric columns
                    # for col in avg_cols:
                    #     if col in df_weekly.columns:
                    #         total_row[col] = df_weekly[col].mean(skipna=True)

                    # Override specific columns with fixed values from executive summary
                    # Get most recent week data
                    most_recent_data = week_metrics[most_recent_week_range]
                    total_row['WTD Ordered SLS $'] = float(
                        most_recent_data["Revenue"].replace("$", "").replace(",", "").strip())
                    total_row['TY 2wks Ago SLS %'] = float(most_recent_data["Prior Period %"].rstrip("%").strip()) / 100
                    total_row['YoY VAR SLS'] = float(most_recent_data["YoY %"].rstrip("%").strip()) / 100
                    total_row['WTD Ordered UNITS'] = float(most_recent_data["Last Week Units"].replace(",", "").strip())
                    total_row['TY 2wks Ago UNITS'] = float(most_recent_data["Prior Units %"].rstrip("%").strip()) / 100
                    total_row['YoY VAR UNITS'] = float(most_recent_data["YoY Units %"].rstrip("%").strip()) / 100

                    # YTD from ytd_metrics
                    total_row['YTD Ordered SLS $'] = float(
                        ytd_metrics["YTD Sales"].replace("$", "").replace(",", "").strip())
                    total_row['YoY VAR YTD SLS'] = float(ytd_metrics["YTD YoY Sales %"].rstrip("%").strip()) / 100
                    total_row['YTD Ordered UNITS'] = float(ytd_metrics["YTD Units"].replace(",", "").strip())
                    total_row['YoY VAR YTD UNITS'] = float(ytd_metrics["YTD YoY Units %"].rstrip("%").strip()) / 100

                    # ==================== INVENTORY TOTALS (Dashboard Grand Totals) ====================
                    if inventory_metrics:
                        try:
                            # Sell-Thru % (already there)
                            if inventory_metrics.get("Sell-Thru %"):
                                val = inventory_metrics["Sell-Thru %"].replace("%", "").strip()
                                total_row['Sell-Thru %'] = float(val) / 100 if val.replace('-', '').replace('.',
                                                                                                            '').replace(
                                    '+', '').isdigit() else np.nan

                            # Open PO QTY
                            if inventory_metrics.get("Open PO QTY"):
                                po_clean = clean_numeric_keep_na(pd.Series([inventory_metrics["Open PO QTY"]]))[0]
                                total_row['Open PO QTY'] = po_clean

                            # On Hand INV $
                            if inventory_metrics.get("On Hand INV $"):
                                inv_clean = clean_numeric_keep_na(pd.Series([inventory_metrics["On Hand INV $"]]))[0]
                                total_row['On Hand INV $'] = inv_clean

                            # On Hand Units
                            if inventory_metrics.get("On Hand Units"):
                                units_clean = clean_numeric_keep_na(pd.Series([inventory_metrics["On Hand Units"]]))[0]
                                total_row['On Hand Units'] = units_clean

                            print(
                                "   Inventory dashboard totals added to Total row (Open PO, On Hand $, Units, Sell-Thru)")

                        except Exception as e:
                            print(f"   Could not parse Inventory totals for Total row: {e}")
                    else:
                        total_row['Sell-Thru %'] = np.nan
                        total_row['Open PO QTY'] = np.nan
                        total_row['On Hand INV $'] = np.nan
                        total_row['On Hand Units'] = np.nan
                    # ===================================================================================

                    # ==================== TRAFFIC TOTALS (Glance Views + % variances) ====================
                    if traffic_metrics:
                        try:
                            # Glance Views (raw number)
                            if traffic_metrics.get("Glance Views"):
                                views_clean = clean_numeric_keep_na(pd.Series([traffic_metrics["Glance Views"]]))[0]
                                total_row['Glance Views'] = views_clean

                            # TY 2wks Ago Views (%)
                            if traffic_metrics.get("TY 2wks Ago Views"):
                                prior_str = traffic_metrics["TY 2wks Ago Views"].rstrip("%").strip()
                                total_row['TY 2wks Ago Views'] = float(prior_str) / 100 if prior_str.replace('-',
                                                                                                             '').replace(
                                    '.', '').replace('+', '').isdigit() else np.nan

                            # YoY VAR Views (%)
                            if traffic_metrics.get("YoY VAR Views"):
                                yoy_str = traffic_metrics["YoY VAR Views"].rstrip("%").strip()
                                total_row['YoY VAR Views'] = float(yoy_str) / 100 if yoy_str.replace('-', '').replace(
                                    '.', '').replace('+', '').isdigit() else np.nan

                            print("   Traffic totals (Glance Views + variances) added to Total row")
                        except Exception as e:
                            print(f"   Could not parse Traffic totals for Total row: {e}")
                    else:
                        total_row['Glance Views'] = np.nan
                        total_row['TY 2wks Ago Views'] = np.nan
                        total_row['YoY VAR Views'] = np.nan
                    # ===================================================================================

                    # ==================== CALCULATE AGGREGATED TOTALS (NOT AVERAGE) ====================
                    # Conversion Rate and ASP $ must be calculated from the grand totals, not averaged per-SKU
                    try:
                        total_rev = total_row.get('WTD Ordered SLS $', 0) or 0
                        total_units = total_row.get('WTD Ordered UNITS', 0) or 0
                        total_views = total_row.get('Glance Views', 0) or 0

                        # ASP $ = Total Revenue / Total Units
                        if total_units > 0:
                            total_row['ASP $'] = total_rev / total_units
                        else:
                            total_row['ASP $'] = np.nan

                        # Conversion Rate = Total Units / Total Views (stored as decimal for % formatting)
                        if total_views > 0:
                            total_row['Conversion Rate'] = total_units / total_views
                        else:
                            total_row['Conversion Rate'] = np.nan

                        print(f"   → Aggregated Total ASP $        : ${total_row.get('ASP $', 0):,.2f}")
                        print(f"   → Aggregated Conversion Rate    : {total_row.get('Conversion Rate', 0) * 100 :.1f}%")

                    except Exception as e:
                        print(f"   Could not calculate aggregated ASP/Conversion for Total row: {e}")
                        total_row['ASP $'] = np.nan
                        total_row['Conversion Rate'] = np.nan
                    # ================================================================================

                    # Append total row
                    df_weekly = pd.concat([df_weekly, pd.DataFrame([total_row])], ignore_index=True)

                    # Save with total row
                    df_weekly.to_excel(final_report_path, index=False, engine='openpyxl')

                    # Apply formats to the updated XLSX
                    wb = load_workbook(final_report_path)
                    ws = wb.active

                    col_map = {col: idx + 1 for idx, col in enumerate(df_weekly.columns)}

                    # Accounting format
                    accounting_fmt = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
                    accounting_fmt_2dec = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

                    for col in ['WTD Ordered SLS $', 'YTD Ordered SLS $', 'On Hand INV $']:
                        if col in col_map:
                            letter = ws.cell(row=1, column=col_map[col]).column_letter
                            for r in range(2, ws.max_row + 1):
                                ws[f"{letter}{r}"].number_format = accounting_fmt

                    for col in ['ASP $', 'COGS', 'Margin ($)']:
                        if col in col_map:
                            letter = ws.cell(row=1, column=col_map[col]).column_letter
                            for r in range(2, ws.max_row + 1):
                                ws[f"{letter}{r}"].number_format = accounting_fmt_2dec

                    # Percentage format (0.0%)
                    percent_cols.append('Margin (%)')
                    for col in percent_cols:
                        if col in col_map:
                            letter = ws.cell(row=1, column=col_map[col]).column_letter
                            for r in range(2, ws.max_row + 1):
                                try:
                                    ws[f"{letter}{r}"].number_format = '0.0%'
                                except Exception as e:
                                    print(f"Error setting percentage format for column '{col}' row {r}: {e}")

                    # Number format
                    for col in ['WTD Ordered UNITS', 'YTD Ordered UNITS', 'Open PO QTY', 'On Hand Units',
                                'Glance Views', 'Wks of INV']:
                        if col in col_map:
                            letter = ws.cell(row=1, column=col_map[col]).column_letter
                            for r in range(2, ws.max_row + 1):
                                ws[f"{letter}{r}"].number_format = '#,##0'

                    # Bold headers and red negatives for specified columns
                    red_font = Font(color="FF0000")
                    bold_font = Font(bold=True)

                    specified_cols = [
                        'Model',
                        'SKUs',
                        'Description',
                        'TY 2wks Ago SLS %',
                        'YoY VAR SLS',
                        'TY 2wks Ago UNITS',
                        'YoY VAR UNITS',
                        'YTD Ordered SLS $',
                        'YoY VAR YTD SLS',
                        'YoY VAR YTD UNITS',
                        'TY 2wks Ago Views',
                        'YoY VAR Views'
                    ]

                    for col in specified_cols:
                        if col in col_map:
                            col_idx = col_map[col]
                            letter = ws.cell(1, col_idx).column_letter
                            # Bold entire column (header + values)
                            for r in range(1, ws.max_row + 1):
                                ws.cell(r, col_idx).font = bold_font
                            # Conditional red for negatives ( < 0 )
                            ws.conditional_formatting.add(
                                f'{letter}2:{letter}{ws.max_row}',
                                CellIsRule(operator='lessThan', formula=['0'], font=red_font)
                            )

                    # Auto-fit columns A-D
                    for col in range(1, 5):  # A=1, B=2, C=3, D=4
                        max_length = 0
                        column_letter = ws.cell(1, col).column_letter
                        for row in range(1, ws.max_row + 1):
                            cell = ws.cell(row, col)
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws.column_dimensions[column_letter].width = max_length * 1.2  # Buffer for readability

                    # Set E+ to width 11 and wrap header text
                    for col in range(5, ws.max_column + 1):
                        column_letter = ws.cell(1, col).column_letter
                        ws.column_dimensions[column_letter].width = 11
                        # Wrap header text
                        ws.cell(1, col).alignment = Alignment(wrap_text=True, vertical='center',
                                                              horizontal='center')

                    # ====================== FREEZE PANES AT F2 ======================
                    ws.freeze_panes = 'F2'
                    print("   Panes frozen at F2 (headers + first 5 columns locked)")
                    # ===============================================================

                    wb.save(final_report_path)
                    print(f"   ✅ Saved {final_report_name} with total row and formats after NetSuite update")

                    # === GENERATE Amazon_YTD_Sales_Report_Revenue.xlsx ===
                    print("\n" + "=" * 90)
                    print("          GENERATING Amazon_YTD_Sales_Report_Revenue.xlsx")
                    print("=" * 90)

                    df_ytd_report_rev = pd.read_excel(final_report_path, dtype=str)
                    df_ytd_report_rev = df_ytd_report_rev[df_ytd_report_rev['Model'] != 'Total']  # Drop the "Total" row
                    df_ytd_report_rev['YTD Ordered SLS $'] = clean_numeric_keep_na(df_ytd_report_rev['YTD Ordered SLS $'])
                    df_ytd_report_rev  = df_ytd_report_rev.dropna(subset=['YTD Ordered SLS $'])
                    df_ytd_report_rev = df_ytd_report_rev.sort_values('YTD Ordered SLS $', ascending=False).head(50)

                    # Add Rank column (1 = highest sales)
                    df_ytd_report_rev.insert(0, 'Rank', range(1, len(df_ytd_report_rev) + 1))

                    ytd_columns_rev = ['Rank', 'Model', 'SKUs', 'Description', 'YTD Ordered SLS $', 'YoY VAR YTD SLS']
                    df_ytd_report_rev = df_ytd_report_rev[ytd_columns_rev]

                    ytd_report_path_rev = os.path.join(FINAL_DATA_DIR, "Amazon_YTD_Sales_Report_Revenue.xlsx")
                    df_ytd_report_rev.to_excel(ytd_report_path_rev, index=False, engine='openpyxl')

                    # Apply formats to new file
                    wb_ytd_rev = load_workbook(ytd_report_path_rev)
                    ws_ytd_rev = wb_ytd_rev.active

                    col_map_ytd_rev = {col: idx + 1 for idx, col in enumerate(ytd_columns_rev)}

                    # Accounting for 'YTD Ordered SLS $'
                    if 'YTD Ordered SLS $' in col_map_ytd_rev:
                        letter = ws_ytd_rev.cell(1, col_map_ytd_rev['YTD Ordered SLS $']).column_letter
                        for r in range(2, ws_ytd_rev.max_row + 1):
                            ws_ytd_rev[f"{letter}{r}"].number_format = accounting_fmt

                    # Percentage for 'YoY VAR YTD SLS'
                    if 'YoY VAR YTD SLS' in col_map_ytd_rev:
                        letter = ws_ytd_rev.cell(1, col_map_ytd_rev['YoY VAR YTD SLS']).column_letter
                        for r in range(2, ws_ytd_rev.max_row + 1):
                            ws_ytd_rev[f"{letter}{r}"].number_format = '0.0%'

                    # Number for 'Rank'
                    if 'Rank' in col_map_ytd_rev:
                        letter = ws_ytd_rev.cell(1, col_map_ytd_rev['Rank']).column_letter
                        for r in range(2, ws_ytd_rev.max_row + 1):
                            ws_ytd_rev[f"{letter}{r}"].number_format = '#,##0'

                    # Auto-fit columns
                    for col in range(1, ws_ytd_rev.max_column + 1):
                        max_length = 0
                        column_letter = ws_ytd_rev.cell(1, col).column_letter
                        for row in range(1, ws_ytd_rev.max_row + 1):
                            cell = ws_ytd_rev.cell(row, col)
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws_ytd_rev.column_dimensions[column_letter].width = max_length * 1.2

                    wb_ytd_rev.save(ytd_report_path_rev)
                    print(f"   ✅ Saved {os.path.basename(ytd_report_path_rev)}")



                    # === GENERATE Amazon_YTD_Sales_Report_Units.xlsx ===
                    print("\n" + "=" * 90)
                    print("          GENERATING Amazon_YTD_Sales_Report_Units.xlsx")
                    print("=" * 90)

                    df_ytd_report_units = pd.read_excel(final_report_path, dtype=str)
                    df_ytd_report_units = df_ytd_report_units[df_ytd_report_units['Model'] != 'Total']  # Drop the "Total" row
                    df_ytd_report_units['YTD Ordered UNITS'] = clean_numeric_keep_na(df_ytd_report_units['YTD Ordered UNITS'])
                    df_ytd_report_units = df_ytd_report_units.dropna(subset=['YTD Ordered UNITS'])
                    df_ytd_report_units = df_ytd_report_units.sort_values('YTD Ordered UNITS', ascending=False).head(50)

                    # Add Rank column (1 = highest sales)
                    df_ytd_report_units.insert(0, 'Rank', range(1, len(df_ytd_report_units) + 1))

                    ytd_columns_units = ['Rank', 'Model', 'SKUs', 'Description', 'YTD Ordered UNITS', 'YoY VAR YTD UNITS']
                    df_ytd_report_units = df_ytd_report_units[ytd_columns_units]

                    ytd_report_path_units = os.path.join(FINAL_DATA_DIR, "Amazon_YTD_Sales_Report_Units.xlsx")
                    df_ytd_report_units.to_excel(ytd_report_path_units, index=False, engine='openpyxl')

                    # Apply formats to new file
                    wb_ytd_units = load_workbook(ytd_report_path_units)
                    ws_ytd_units = wb_ytd_units.active

                    col_map_ytd_units = {col: idx + 1 for idx, col in enumerate(ytd_columns_units)}

                    # Number for 'YTD Ordered UNITS'
                    if 'YTD Ordered UNITS' in col_map_ytd_units:
                        letter = ws_ytd_units.cell(1, col_map_ytd_units['YTD Ordered UNITS']).column_letter
                        for r in range(2, ws_ytd_units.max_row + 1):
                            ws_ytd_units[f"{letter}{r}"].number_format = '#,##0'

                    # Percentage for 'YoY VAR YTD UNITS'
                    if 'YoY VAR YTD UNITS' in col_map_ytd_units:
                        letter = ws_ytd_units.cell(1, col_map_ytd_units['YoY VAR YTD UNITS']).column_letter
                        for r in range(2, ws_ytd_units.max_row + 1):
                            ws_ytd_units[f"{letter}{r}"].number_format = '0.0%'

                    # Number for 'Rank'
                    if 'Rank' in col_map_ytd_units:
                        letter = ws_ytd_units.cell(1, col_map_ytd_units['Rank']).column_letter
                        for r in range(2, ws_ytd_units.max_row + 1):
                            ws_ytd_units[f"{letter}{r}"].number_format = '#,##0'

                    # Auto-fit columns
                    for col in range(1, ws_ytd_units.max_column + 1):
                        max_length = 0
                        column_letter = ws_ytd_units.cell(1, col).column_letter
                        for row in range(1, ws_ytd_units.max_row + 1):
                            cell = ws_ytd_units.cell(row, col)
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws_ytd_units.column_dimensions[column_letter].width = max_length * 1.2

                    wb_ytd_units.save(ytd_report_path_units)
                    print(f"   ✅ Saved {os.path.basename(ytd_report_path_units)}")

                    # === CAPTURE EXECUTIVE SUMMARY AS HTML ===
                    exec_summary_html = f"""
                    <h1>AMAZON SALES EXECUTIVE SUMMARY</h1>

                    <h4>Weekly Data Summary</h4>
                    <table border="1" cellspacing="0" cellpadding="5">
                    <tr>
                        <th style="text-align: left; background-color: #f2f2f2;">Week</th>
                        <th style="text-align: right; background-color: #f2f2f2;">Revenue</th>
                        <th style="text-align: right; background-color: #f2f2f2;">vs Prior Week</th>
                        <th style="text-align: right; background-color: #f2f2f2;">vs Same Week - Last Year</th>
                        <th style="text-align: right; background-color: #f2f2f2;">Total Units Sold</th>
                    </tr>
                    """

                    sorted_weeks = sorted(week_metrics.items(), key=lambda x: x[1]["Week Start"], reverse=True)

                    revenues = []
                    for week_range, data in sorted_weeks:
                        try:
                            rev_clean = data["Revenue"].replace("$", "").replace(",", "").strip()
                            rev_float = float(rev_clean)
                            rev_rounded = round(rev_float)
                            revenues.append(rev_float)
                        except:
                            rev_rounded = "ERROR"
                            revenues.append(0)

                        prior_str = data["Prior Period %"].rstrip("%").strip()
                        prior_val = float(prior_str) if prior_str.replace('-', '').replace('.', '').replace('+',
                                                                                                            '').isdigit() else 0
                        prior_color = 'color: red;' if prior_val < 0 else ''
                        prior_str = f"{prior_val:+.1f}%" if prior_str.replace('-', '').replace('.', '').replace('+',
                                                                                                                '').isdigit() else prior_str
                        prior_html = f'<td style="text-align: right; font-weight: bold; {prior_color}">{prior_str}</td>'

                        yoy_str = data["YoY %"].rstrip("%").strip()
                        yoy_val = float(yoy_str) if yoy_str.replace('-', '').replace('.', '').replace('+',
                                                                                                      '').isdigit() else 0
                        yoy_color = 'color: red;' if yoy_val < 0 else ''
                        yoy_str = f"{yoy_val:+.2f}%" if yoy_str.replace('-', '').replace('.', '').replace('+',
                                                                                                          '').isdigit() else yoy_str
                        yoy_html = f'<td style="text-align: right; font-weight: bold; {yoy_color}">{yoy_str}</td>'

                        units_raw = data["Last Week Units"].replace(",", "").strip()
                        units = f"{int(units_raw):,}" if units_raw.isdigit() else "N/A"
                        units_html = f'<td style="text-align: right; font-weight: bold;">{units}</td>'

                        start_dt = datetime.strptime(data["Week Start"], "%Y-%m-%d").date()
                        end_dt = datetime.strptime(data["Week End"], "%Y-%m-%d").date()

                        week_info = get_business_week_info(start_dt)
                        week_num = week_info["week_number"]
                        week_label = f"Week {week_num:02d} | {start_dt:%Y-%m-%d} to {end_dt:%Y-%m-%d}"
                        week_html = f'<td style="text-align: left; font-weight: bold;">{week_label}</td>'

                        rev_html = f'<td style="text-align: right; font-weight: bold;">${rev_rounded:,.0f}</td>'

                        exec_summary_html += f"<tr>{week_html} {rev_html} {prior_html} {yoy_html} {units_html}</tr>"

                    exec_summary_html += "</table>"

                    exec_summary_html += "<br><br>"

                    if len(revenues) == 4 and sum(revenues) > 0:
                        four_week_avg = sum(revenues) / 4
                        four_week_avg_html = f'<td style="text-align: right; font-weight: bold;">${round(four_week_avg):,.0f}</td>'

                        most_recent = revenues[0]
                        most_recent_html = f'<td style="text-align: right; font-weight: bold;">${round(most_recent):,.0f}</td>'

                        growth_vs_avg = (most_recent / four_week_avg - 1) * 100
                        growth_color = 'color: red;' if growth_vs_avg < 0 else ''
                        growth_html = f'<td style="text-align: right; font-weight: bold; {growth_color}">{growth_vs_avg:+.2f}%</td>'


                        exec_summary_html += """
                        <table border="1" cellspacing="0" cellpadding="5">
                        <tr>
                            <th style="text-align: left; background-color: #f2f2f2;">Metric</th>
                            <th style="text-align: right; background-color: #f2f2f2;">Value</th>
                        </tr>
                        """
                        exec_summary_html += f"<tr><td style='text-align: left; font-weight: bold;'>Most Recent Week Sales</td> {most_recent_html}</tr>"
                        exec_summary_html += f"<tr><td style='text-align: left; font-weight: bold;'>4-Week Average Sales</td> {four_week_avg_html}</tr>"
                        exec_summary_html += f"<tr><td style='text-align: left; font-weight: bold;'>Growth vs 4-Week Average</td> {growth_html}</tr>"
                        exec_summary_html += "</table>"
                    else:
                        exec_summary_html += "<h4>4-Week Summary</h4>"
                        exec_summary_html += """
                        <table border="1" cellspacing="0" cellpadding="5">
                        <tr>
                            <th style="text-align: left; background-color: #f2f2f2;">Metric</th>
                            <th style="text-align: right; background-color: #f2f2f2;">Value</th>
                        </tr>
                        """
                        exec_summary_html += "<tr><td colspan='2'>Could not calculate summary — missing data</td></tr>"
                        exec_summary_html += "</table>"

                    exec_summary_html += f"""
                    <h4>{ytd_metrics.get('Period', 'YEAR-TO-DATE')}</h4>
                    <table border="1" cellspacing="0" cellpadding="5">
                    <tr>
                        <th style="text-align: left; background-color: #f2f2f2;">Metric</th>
                        <th style="text-align: right; background-color: #f2f2f2;">Value</th>
                    </tr>
                    """

                    if "YTD Sales" in ytd_metrics:
                        try:
                            ytd_sales_clean = ytd_metrics["YTD Sales"].replace("$", "").replace(",", "").strip()
                            ytd_sales_float = float(ytd_sales_clean)
                            ytd_sales_rounded = round(ytd_sales_float)
                            ytd_sales_html = f'<td style="text-align: right; font-weight: bold;">${ytd_sales_rounded:,.0f}</td>'

                            ytd_yoy_sales = ytd_metrics["YTD YoY Sales %"].rstrip("%").strip()
                            yoy_sales_val = float(ytd_yoy_sales)
                            yoy_sales_color = 'color: red;' if yoy_sales_val < 0 else ''
                            yoy_sales_str = f"{yoy_sales_val:+.2f}%"
                            yoy_sales_html = f'<td style="text-align: right; font-weight: bold; {yoy_sales_color}">{yoy_sales_str}</td>'

                            ytd_units_clean = ytd_metrics["YTD Units"].replace(",", "").strip()
                            ytd_units_int = int(float(ytd_units_clean))
                            ytd_units_html = f'<td style="text-align: right; font-weight: bold;">{ytd_units_int:,.0f}</td>'

                            ytd_yoy_units = ytd_metrics["YTD YoY Units %"].rstrip("%").strip()
                            yoy_units_val = float(ytd_yoy_units)
                            yoy_units_color = 'color: red;' if yoy_units_val < 0 else ''
                            yoy_units_str = f"{yoy_units_val:+.2f}%"
                            yoy_units_html = f'<td style="text-align: right; font-weight: bold; {yoy_units_color}">{yoy_units_str}</td>'

                            exec_summary_html += f"<tr><td style='text-align: left; font-weight: bold;'>YTD Ordered Sales</td> {ytd_sales_html}</tr>"
                            exec_summary_html += f"<tr><td style='text-align: left; font-weight: bold;'>YTD YoY Sales Growth</td> {yoy_sales_html}</tr>"
                            exec_summary_html += f"<tr><td style='text-align: left; font-weight: bold;'>YTD Ordered Units</td> {ytd_units_html}</tr>"
                            exec_summary_html += f"<tr><td style='text-align: left; font-weight: bold;'>YTD YoY Units Growth</td> {yoy_units_html}</tr>"

                        except:
                            exec_summary_html += "<tr><td colspan='2'>Could not parse YTD numbers</td></tr>"

                    exec_summary_html += "</table>"

                    if business_metrics:
                        try:
                            bus_rev_clean = business_metrics["Revenue"].replace("$", "").replace(",", "").strip()
                            bus_rev_float = float(bus_rev_clean)
                            bus_rev_rounded = round(bus_rev_float)
                            bus_rev_html = f'<td style="text-align: right; font-weight: bold;">${bus_rev_rounded:,.0f}</td>'

                            bus_yoy = business_metrics["YoY"].rstrip("%").strip()
                            bus_yoy_val = float(bus_yoy)
                            bus_yoy_color = 'color: red;' if bus_yoy_val < 0 else ''
                            bus_yoy_str = f"{bus_yoy_val:+.2f}%"
                            bus_yoy_html = f'<td style="text-align: right; font-weight: bold; {bus_yoy_color}">{bus_yoy_str}</td>'

                            latest_rev_clean = week_metrics[most_recent_week_range]["Revenue"].replace("$", "").replace(
                                ",", "").strip()
                            latest_rev_float = float(latest_rev_clean)
                            bus_pct = (bus_rev_float / latest_rev_float) * 100 if latest_rev_float != 0 else 0
                            bus_pct_str = f"{bus_pct:.2f}%"
                            bus_pct_html = f'<td style="text-align: right; font-weight: bold;">{bus_pct_str}</td>'

                            exec_summary_html += "<h4>Amazon Business</h4>"
                            exec_summary_html += """
                            <table border="1" cellspacing="0" cellpadding="5">
                            <tr>
                                <th style="text-align: left; background-color: #f2f2f2;">Metric</th>
                                <th style="text-align: right; background-color: #f2f2f2;">Value</th>
                            </tr>
                            """
                            exec_summary_html += f"<tr><td style='text-align: left; font-weight: bold;'>Business Revenue</td> {bus_rev_html}</tr>"
                            exec_summary_html += f"<tr><td style='text-align: left; font-weight: bold;'>vs Same Week - Last Year</td> {bus_yoy_html}</tr>"
                            exec_summary_html += f"<tr><td style='text-align: left; font-weight: bold;'>Business % of Total Sales</td> {bus_pct_html}</tr>"
                            exec_summary_html += "</table>"
                        except:
                            exec_summary_html += "<h4>Amazon Business</h4><table border='1' cellspacing='0' cellpadding='5'><tr><td colspan='2'>Could not parse Business metrics</td></tr></table>"
                    else:
                        exec_summary_html += "<h4>Amazon Business</h4><table border='1' cellspacing='0' cellpadding='5'><tr><td colspan='2'>Business metrics not available</td></tr></table>"

                    def df_to_styled_html(df, is_ytd=False):
                        # Define columns where negative values should be red and all values bold
                        red_negative_and_bold_columns = [
                            'TY 2wks Ago SLS %',
                            'YoY VAR SLS',
                            'TY 2wks Ago UNITS',
                            'YoY VAR UNITS',
                            'YoY VAR YTD SLS',
                            'YoY VAR YTD UNITS',
                            'TY 2wks Ago Views',
                            'YoY VAR Views'
                        ]

                        # Define columns where all values should be bold
                        bold_columns = [
                            'On Hand INV $',
                            'On Hand Units',
                            'Wks of INV',
                            'Sell-Thru %',
                            'Glance Views',
                            'Conversion Rate',
                            'ASP $',
                            'Model',
                            'SKUs',
                            'Description',
                            'Category',
                            'ASIN',
                            'WTD Ordered SLS $',
                            'WTD Ordered UNITS',
                            'YTD Ordered SLS $',
                            'YTD Ordered UNITS',
                            'Open PO QTY',
                            'COGS',
                            'Margin ($)',
                            'Margin (%)'
                        ]

                        html = '<table border="1" cellspacing="0" cellpadding="5">'
                        html += '<tr>'
                        for col in df.columns:
                            html += f'<th style="text-align: center; background-color: #f2f2f2; word-wrap: break-word;">{col}</th>'
                        html += '</tr>'

                        for _, row in df.iterrows():
                            html += '<tr>'
                            for col_idx, col in enumerate(df.columns):
                                raw_val = row[col]
                                # ── This is the ONLY place we convert NaN → empty string ──
                                display_val = format_for_display(raw_val, col_name=col)

                                style = f'text-align: {"left" if col_idx < 3 else "right"};'
                                if pd.isna(raw_val):
                                    style += 'color: #888888;'  # optional: light gray for empty cells

                                # Apply bold to all values in specified columns
                                if col in bold_columns or col in red_negative_and_bold_columns:
                                    style += 'font-weight: bold;'

                                # Apply red to negative numeric values in specified columns
                                if col in red_negative_and_bold_columns and isinstance(raw_val,
                                                                                       (int, float)) and raw_val < 0:
                                    style += 'color: red;'

                                # Prevent wrapping for columns after the first three
                                if col_idx >= 3:
                                    style += 'white-space: nowrap;'

                                html += f'<td style="{style}">{display_val}</td>'
                            html += '</tr>'

                        html += '</table>'
                        return html

                    # === CREATE OUTLOOK EMAIL ===
                    print("\n" + "=" * 90)
                    print("          CREATING OUTLOOK EMAIL")
                    print("=" * 90)

                    # For weekly top 40 + total
                    df_weekly_full = pd.read_excel(final_report_path)
                    for col in df_weekly_full.columns:
                        if col in cols_to_clean:
                            df_weekly_full[col] = clean_numeric_keep_na(df_weekly_full[col])
                    df_weekly_top40 = pd.concat([df_weekly_full[:-1].head(40), df_weekly_full.tail(1)])

                    # Define snapshot columns (adjust as needed; this drops less essential ones like percentages/variances)
                    snapshot_columns = ['Model', 'SKUs', 'Description', 'Category', 'ASIN', 'WTD Ordered SLS $',
                                        'WTD Ordered UNITS', 'YTD Ordered SLS $', 'YTD Ordered UNITS', 'YoY VAR YTD SLS',
                                        'Open PO QTY', 'On Hand INV $', 'On Hand Units', 'Wks of INV', 'Sell-Thru %',
                                        'Glance Views', 'Conversion Rate', 'ASP $'
                                        # , 'COGS', 'Margin ($)', 'Margin (%)' Dropped until COGS data in NetSuite is fixed
                                        ]

                    df_weekly_snapshot = df_weekly_top40[snapshot_columns]

                    weekly_table_html = df_to_styled_html(df_weekly_snapshot)

                    # For YTD Revenue
                    df_ytd_rev = pd.read_excel(ytd_report_path_rev)
                    for col in df_ytd_rev.columns:
                        if col in cols_to_clean:
                            df_ytd_rev[col] = clean_numeric_keep_na(df_ytd_rev[col])
                    ytd_rev_table_html = df_to_styled_html(df_ytd_rev, is_ytd=True)

                    # For YTD Units
                    df_ytd_units = pd.read_excel(ytd_report_path_units)
                    for col in df_ytd_units.columns:
                        if col in cols_to_clean:
                            df_ytd_units[col] = clean_numeric_keep_na(df_ytd_units[col])
                    ytd_units_table_html = df_to_styled_html(df_ytd_units, is_ytd=True)

                    # Create email
                    outlook = win32.Dispatch('outlook.application')
                    mail = outlook.CreateItem(0)  # 0: olMailItem
                    mail.Subject = f'Amazon Sales Report - {year_for_email} - Week {week_num_for_email:02d}'
                    mail.To = os.environ.get('REPORT_EMAIL_TO', 'your-email@your-company.com')  # Replace with actual email
                    mail.CC = ''  # Optional
                    mail.BodyFormat = 2  # olFormatHTML
                    mail.HTMLBody = f"""
                    <html>
                    <body>
                    <img src="{os.path.join(os.environ.get("SALES_DATA_DIR", "./data"), "PNG Files", "amazon-logo.png")}" 
                     alt="Amazon Logo" 
                     width="800" 
                     style="width:800px; max-width:800px; height:auto; margin-bottom:60px; display:block;"><br>
                    <br>
                    <br>
                    {exec_summary_html}
                    <h2>Weekly Top 40 SKUs Snapshot</h2>
                    {weekly_table_html}
                    <h2>YTD Sales Report - Top 50 SKUs</h2>
                    <div style="text-align: center;">
                    <table style="margin: 0 auto;">
                    <tr>
                        <td style="text-align: center;">By Revenue</td>
                        <td style="text-align: center;">By Units</td>
                    </tr>
                    <tr>
                        <td style="vertical-align: top;">{ytd_rev_table_html}</td>
                        <td style="vertical-align: top;">{ytd_units_table_html}</td>
                    </tr>
                    </table>
                    <h2>YoY Weekly Revenue</h2>
                    <img src="{resized_img_path}" alt="YOY Line Graph" style="border: 2px solid #000000; display: block;">
                    <h2>Sales by Category (Top 5)</h2>
                    <div style="text-align: center;">
                    <table style="margin: 0 auto;">                        
                    <tr>
                        <td><img src="{week_rev_pie_img_path}" alt="WTD Sales by Category ($) Pie Chart" style="border: 2px solid #000000; display: block;"></td>
                        <td><img src="{ytd_rev_pie_img_path}" alt="YTD Sales by Category ($) Pie Chart" style="border: 2px solid #000000; display: block;"></td>
                    </tr>
                    <tr>
                        <td><img src="{week_units_pie_img_path}" alt="WTD Sales by Category (Units) Pie Chart" style="border: 2px solid #000000; display: block;"></td>
                        <td><img src="{ytd_units_pie_img_path}" alt="YTD Sales by Category (Units) Pie Chart" style="border: 2px solid #000000; display: block;"></td>
                    </tr>
                    </table>
                    </div>
                    <h2>Category Trends</h2>
                    <div style="text-align: center;">
                    <table style="margin: 0 auto;">
                    <tr>
                        <td><img src="{category_rev_line_path}" alt="Weekly Revenue by Category Line Graph" style="border: 2px solid #000000; display: block;"></td>
                        <td><img src="{category_units_line_path}" alt="Weekly Units by Category Line Graph" style="border: 2px solid #000000; display: block;"></td>
                    </tr>
                    </table>
                    </div>
                    </body>
                    </html>
                    """

                    # Removed cumalitive data from above. Charts are still generated behind the scenes if fixed in the future


                    # Attach files if needed (e.g., full reports)
                    mail.Attachments.Add(final_report_path)

                    # Display the email (ready to send)
                    mail.Display()
                    # mail.Send()

                    # Clean up temp image after sending? User can send manually.
                    print("   Outlook email drafted and displayed.")

                else:
                    print("   No ASIN column — skipping final step")

            end_time = time.time()
            duration = end_time - start_time
            minutes = int(duration // 60)
            seconds = int(duration % 60)
            print(f"\nReport generation completed in {minutes} minutes {seconds} seconds")

    except Exception as e:
        log(f"Script failed: {e}")
        driver.save_screenshot("final_error.png")

    finally:
        # driver.quit()  # Uncomment for automated runs
        pass

if __name__ == "__main__":
    main()


    # 1898